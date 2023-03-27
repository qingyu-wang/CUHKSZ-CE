"""
openpyxl number_format
https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
"""

import datetime
import traceback

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, numbers

from .utils_mongo import mongo


def get_worksheet_value_by_rows(ws):
    values = []
    for row in ws.iter_rows():
        values.append([cell.value for cell in row])
    return values


def preprocess_zoom(old_wb_path, new_wb_path):
    result = {
        "msgs": [],
    }

    # 读取文件
    try:
        # 读取原始数据
        old_wb = load_workbook(old_wb_path)
        old_ws = old_wb.worksheets[0]
        old_raw_values = get_worksheet_value_by_rows(old_ws)
        old_titles = old_raw_values[0]
        old_values = old_raw_values[1:]

        # 清洗数据
        infos = {}
        for idx, row in enumerate(old_values):
            # 核对是否为访客
            if row[5] == "是":
                campus_idno = "访客 (名称: %s)" % row[0]
            else:
                # 核对是否为学生
                if "@link.cuhk.edu.cn" in row[1]:
                    campus_idno = row[1].split("@link.cuhk.edu.cn")[0]
                    find_student_filter = {"campus_role": "学生", "campus_idno": campus_idno}
                    find_student_num = mongo.coll_user_info.count_documents(find_student_filter)
                    if find_student_num != 1:
                        campus_idno = "学生 (查询结果错误: %s, 邮件: %s)" % (find_student_num, row[1])
                else:
                    continue
                    # campus_idno = "教工 (邮件: %s)" % row[1]
            # 签到记录
            time_signin  = row[2]
            time_signout = row[3]
            if isinstance(time_signin, str):
                time_signin  = datetime.datetime.strptime(time_signin,  "%m/%d/%Y %I:%M:%S %p")
            if isinstance(time_signout, str):
                time_signout = datetime.datetime.strptime(time_signout, "%m/%d/%Y %I:%M:%S %p")
            assert isinstance(time_signin,  datetime.datetime)
            assert isinstance(time_signout, datetime.datetime)
            # 持续时间
            duration = time_signout - time_signin
            # 更新信息
            if campus_idno not in infos:
                infos[campus_idno] = {
                    "校园卡号": campus_idno,
                    "持续时间": duration,
                    "签到记录": [time_signin],
                    "签退记录": [time_signout]
                }
            else:
                infos[campus_idno]["持续时间"] += duration
                infos[campus_idno]["签到记录"].append(time_signin)
                infos[campus_idno]["签退记录"].append(time_signout)

        # 重新排序
        infos_temp_1 = []
        infos_temp_2 = []
        for info in infos.values():
            if info["校园卡号"][:2] in ["访客", "学生", "教工"]:
                infos_temp_1.append(info)
            else:
                infos_temp_2.append(info)
        infos_sort = []
        infos_sort += sorted(infos_temp_1, key=lambda x: x["持续时间"], reverse=True)
        infos_sort += sorted(infos_temp_2, key=lambda x: x["持续时间"], reverse=True)

    except Exception as e:
        print(traceback.format_exc())
        result["msgs"].append({
            "type": "error",
            "text": "读取文件失败<br>[读取文件] <= \"%s\"<br>[保存文件] => \"%s\"" % (
                old_wb_path, new_wb_path
            )
        })
        return result

    # 保存文件
    try:
        new_wb = Workbook()
        new_ws = new_wb.worksheets[0]

        headers = ["校园卡号", "持续时间 (秒)", "持续时间", "签到记录", "签退记录"]

        new_aligntment = {
            "校园卡号": Alignment(
                vertical="center",
                # horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "持续时间 (秒)": Alignment(
                vertical="center",
                horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "持续时间": Alignment(
                vertical="center",
                horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "签到记录": Alignment(
                vertical="center",
                # horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "签退记录": Alignment(
                vertical="center",
                # horizontal="center",
                # wrap_text=True # 自动换行
            ),
        }
        new_font = {
            "header": Font(
                bold=True
            )
        }

        new_ws.column_dimensions["A"].width = 20
        new_ws.column_dimensions["B"].width = 20
        new_ws.column_dimensions["C"].width = 20
        new_ws.column_dimensions["D"].width = 20
        new_ws.column_dimensions["E"].width = 20

        for col, key in enumerate(headers, start=1):
            cell = new_ws.cell(row=1, column=col)
            cell.value = key
            cell.alignment = new_aligntment[key]
            cell.font = new_font["header"]
            cell.number_format = numbers.FORMAT_TEXT

        for row, info in enumerate(infos_sort, start=2):
            for col, key in enumerate(headers, start=1):
                cell = new_ws.cell(row=row, column=col)
                if key == "签到记录":
                    cell.value = "\n".join(sorted([j.strftime("%Y-%m-%d %H:%M:%S") for j in info[key]]))
                    cell.number_format = numbers.FORMAT_TEXT
                elif key == "签退记录":
                    cell.value = "\n".join(sorted([j.strftime("%Y-%m-%d %H:%M:%S") for j in info[key]]))
                    cell.number_format = numbers.FORMAT_TEXT
                elif key == "持续时间 (秒)":
                    cell.value = info["持续时间"].total_seconds()
                    cell.number_format = numbers.FORMAT_NUMBER
                elif key == "持续时间":
                    cell.value = str(info[key])
                    cell.number_format = numbers.FORMAT_TEXT
                else:
                    cell.value = info[key]
                    cell.number_format = numbers.FORMAT_TEXT
                cell.alignment = new_aligntment[key]

        new_wb.save(new_wb_path)

    except Exception as e:
        print(traceback.format_exc())
        result["msgs"].append({
            "type": "error",
            "text": "保存文件失败<br>[读取文件] <= \"%s\"<br>[保存文件] => \"%s\"" % (
                old_wb_path, new_wb_path
            )
        })
        return result

    result["msgs"].append({
        "type": "success",
        "text": "处理成功"
    })
    return result


if __name__ == "__main__":

    """
python -m utils.utils_tool_zoom
    """

    old_wb_path = "./data/test/tool_utils-zoom-unprocessed.xlsx"
    new_wb_path = "./data/test/tool_utils-zoom.xlsx"

    result = preprocess_zoom(old_wb_path, new_wb_path)
    print(result["msgs"][-1]["type"])
    print("[读取文件] %s" % old_wb_path)
    print("[保存文件] %s" % new_wb_path)
