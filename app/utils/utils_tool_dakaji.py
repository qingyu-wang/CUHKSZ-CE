# -*- coding: utf-8- *-

"""
openpyxl number_format
https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
"""

import datetime
import traceback

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

from .utils_oracle import oracle


def search_dakaji(search_text):
    query = """
    SELECT
        c.conferencename    AS "name",
        c.conferencedate    AS "date",
        c.starttime         AS "stime",
        c.endtime           AS "etime"
    FROM
        conference c
    WHERE
        c.conferencename LIKE '%%%s%%'
    ORDER BY
        c.conferencedate DESC
    """
    search_infos = []
    oracle.execute(query % search_text)
    headers = [i[0] for i in oracle.cursor.description]
    for row in oracle.cursor:
        search_info = dict(zip(headers, row))
        search_infos.append(search_info)
    return search_infos


def preprocess_dakaji(search_text, save_path):

    result = {
        "msgs": [],
    }

    query = """
    SELECT
        t.account           AS "ID (打卡系统)",
        trim(t.name)        AS "姓名",
        t.sex               AS "性别",
        trim(t.identity)    AS "身份证号",
        trim(t.sno)         AS "校园卡号",
        t.p_code            AS "校园身份ID",
        t.p_name            AS "校园身份名称",
        t.d1_code           AS "一级部门ID",
        t.d1_name           AS "一级部门名称",
        t.d2_code           AS "二级部门ID",
        t.d2_name           AS "二级部门名称",
        c.conferencename    AS "会议名称",
        c.conferencedate    AS "会议日期",
        c.starttime         AS "开始时间",
        c.endtime           AS "结束时间",
        cw.watertime        AS "打卡时间"
    FROM
        conferencewaterlist cw
        INNER JOIN conference c ON c.id = cw.conferenceid
        INNER JOIN (
            SELECT
                a.account,
                a.name,
                a.sex,
                a.identity,
                a.sno,
                p.code AS p_code,
                p.name AS p_name,
                d.d1_code,
                d.d1_name,
                d.d2_code,
                d.d2_name
            FROM
                iddbuser.account a
                INNER JOIN iddbuser.pid p ON p.code = a.pid
                INNER JOIN (
                    SELECT
                        d1.code AS d1_code,
                        d1.name AS d1_name,
                        d2.code AS d2_code,
                        d2.name AS d2_name
                    FROM
                        iddbuser.department d2
                        INNER JOIN iddbuser.department d1 ON d1.code = substr(d2.code, 1, 3)
                ) d ON d.d2_code = trim(substr(a.deptcode, 4))
        ) t ON t.account = cw.accountid
    WHERE
        c.conferencename = '%s'
    ORDER BY
        cw.watertime,
        t.sno
    """

    # 查询数据
    try:
        # 读取原始数据
        rows = [row for row in oracle.cursor.execute(query % search_text)]

        if not rows:
            result["msgs"].append({
                "type": "warn",
                "text": "未查询到结果<br>[ 查询字段=\"%s\" ]" % search_text
            })
            return result

        # 清洗数据
        infos = {}
        for idx, row in enumerate(rows):
            name = row[1]
            campus_idno = row[4]
            dept = row[8]
            time_signin = datetime.datetime.strptime(row[15], "%Y-%m-%d %H:%M:%S")
            if dept != "学生":
                continue
            else:
                if campus_idno not in infos:
                    infos[campus_idno] = {
                        "校园卡号": campus_idno,
                        "持续时间": datetime.timedelta(0),
                        "签到记录": [time_signin],
                    }
                else:
                    infos[campus_idno]["签到记录"].append(time_signin)
                    infos[campus_idno]["持续时间"] = infos[campus_idno]["签到记录"][-1] - infos[campus_idno]["签到记录"][0]

        # 重新排序
        infos_sort = sorted(infos.values(), key=lambda x: x["校园卡号"])
        infos_sort = sorted(infos_sort, key=lambda x: x["持续时间"], reverse=True)

    except Exception as e:
        print(traceback.format_exc())
        result["msgs"].append({
            "type": "error",
            "text": "查询数据失败<br>[查询字段] <= \"%s\"<br>[保存文件] => \"%s\"" % (
                search_text, save_path
            )
        })
        return result

    # 保存文件
    try:
        new_wb = Workbook()
        new_ws = new_wb.worksheets[0]

        headers = ["校园卡号", "持续时间 (秒)", "持续时间", "签到记录"]

        new_aligntment = {
            "校园卡号": Alignment(
                vertical="center",
                # horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "持续时间 (秒)": Alignment(
                vertical="center",
                horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "持续时间": Alignment(
                vertical="center",
                horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "签到记录": Alignment(
                vertical="center",
                # horizontal="center",
                # wrap_text=True # 自动换行
            ),
        }
        new_font = {
            "header": Font(
                bold=True
            )
        }

        new_ws.column_dimensions["A"].width = 20
        new_ws.column_dimensions["B"].width = 20
        new_ws.column_dimensions["C"].width = 20
        new_ws.column_dimensions["D"].width = 20

        for col, key in enumerate(headers, start=1):
            cell = new_ws.cell(row=1, column=col)
            cell.value = key
            cell.alignment = new_aligntment[key]
            cell.font = new_font["header"]
            cell.number_format = numbers.FORMAT_TEXT

        for row, info in enumerate(infos_sort, start=2):
            for col, key in enumerate(headers, start=1):
                cell = new_ws.cell(row=row, column=col)
                if key == "签到记录":
                    cell.value = "\n".join([j.strftime("%Y-%m-%d %H:%M:%S") for j in info[key]])
                    cell.number_format = numbers.FORMAT_TEXT
                elif key == "持续时间 (秒)":
                    cell.value = info["持续时间"].total_seconds()
                    cell.number_format = numbers.FORMAT_NUMBER
                elif key == "持续时间":
                    cell.value = str(info[key])
                    cell.number_format = numbers.FORMAT_TEXT
                else:
                    cell.value = info[key]
                    cell.number_format = numbers.FORMAT_TEXT
                cell.alignment = new_aligntment[key]

        new_wb.save(save_path)

    except Exception as e:
        print(traceback.format_exc())
        result["msgs"].append({
            "type": "error",
            "text": "保存文件失败<br>[查询字段] <= \"%s\"<br>[保存文件] => \"%s\"" % (
                search_text, save_path
            )
        })
        return result

    result["msgs"].append({
        "type": "success",
        "text": "处理成功"
    })
    return result


if __name__ == "__main__":

    """
python -m utils.utils_tool_dakaji
    """

    search_text = "20220802-校门准入"
    save_path = "./data/test/tool_utils-dakaji.xlsx"

    result = preprocess_dakaji(search_text, save_path)
    print(result["msgs"][-1]["type"])
    print("[保存文件] %s" % save_path)
