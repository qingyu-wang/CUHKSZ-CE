"""
openpyxl number_format
https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html

问卷星 2 (姓名, 校园卡号)
"""

import traceback

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter

from .utils_mongo import mongo


def get_worksheet_value_by_rows(ws):
    values = []
    for row in ws.iter_rows():
        values.append([cell.value for cell in row])
    return values


def preprocess_wenjuanxing_2(old_wb_path, new_wb_path):
    result = {
        "msgs": [],
    }

    # 读取文件
    try:
        # 读取原始数据
        old_wb = load_workbook(old_wb_path)
        old_ws = old_wb.worksheets[0]
        old_raw_values = get_worksheet_value_by_rows(old_ws)
        old_titles = old_raw_values[0]
        old_values = old_raw_values[1:]

        assert old_titles[0] == "姓名"
        assert old_titles[1] == "校园卡号"

        custom_titles = ["[自定义] %s" % i for i in old_titles[2:]]

        num_rows = old_ws.max_row
        num_cols = old_ws.max_column

        # 清洗数据
        infos = dict() # 储存数据
        campus_idno_pool = dict() # 储存校园卡号，去重
        for idx, value in enumerate(old_values):
            info = dict()
            info["系统备注"] = ""
            campus_idno_duplicated = False

            # 提取字段 (姓名, 校园卡号, 邮箱)
            name_1 = str(value[0]).strip()
            campus_idno_1 = str(value[1]).strip()

            for i, custom_title in enumerate(custom_titles):
                info[custom_title] = value[i+2]

            # 核对校园卡号
            find_student_num_by_idno_1 = mongo.coll_user_info.count_documents({"campus_role": "学生", "campus_idno": campus_idno_1})
            if find_student_num_by_idno_1 > 1:
                print("find student error [ find_student_num_by_idno_1=%s, campus_idno_1=%s ]" % (find_student_num_by_idno_1, campus_idno_1))
                continue
            if find_student_num_by_idno_1 != 0:
                info["校园卡号"] = campus_idno_1
            else:
                info["校园卡号"] = "卡号未知"
                info["系统备注"] += "%s卡号不存在 (%s)" % (
                    "" if len(info["系统备注"]) == 0 else "\n",
                    campus_idno_1
                )

            # 核对姓名
            find_student_num_by_name_1 = mongo.coll_user_info.count_documents({"campus_role": "学生", "name": name_1})
            if info["校园卡号"] == "卡号未知":
                if find_student_num_by_name_1 != 0:
                    info["姓名"] = name_1
                else:
                    info["姓名"] = "姓名未知"
                    info["系统备注"] += "%s姓名不存在 (%s)" % (
                        "" if len(info["系统备注"]) == 0 else "\n",
                        name_1
                        )
                # 根据姓名反查校园卡好
                if info["姓名"] != "姓名未知":
                    find_campus_idnos_by_name = [datum["campus_idno"] for datum in mongo.coll_user_info.find({"campus_role": "学生", "name": info["姓名"]})]
                    if campus_idno_1 in find_campus_idnos_by_name:
                        info["校园卡号"] = campus_idno_1
                    else:
                        info["系统备注"] += "%s校园卡号与姓名不一致 (%s | 系统: %s)" % (
                            "" if len(info["系统备注"]) == 0 else "\n",
                            campus_idno_1,
                            ", ".join(find_campus_idnos_by_name)
                        )
            else:
                find_student_name = mongo.coll_user_info.find_one({"campus_role": "学生", "campus_idno": info["校园卡号"]})["name"]
                info["姓名"] = find_student_name
                if find_student_name not in [name_1]:
                    info["系统备注"] += "%s姓名与校园卡号不一致 (%s)" % (
                        "" if len(info["系统备注"]) == 0 else "\n",
                        name_1
                    )

            info["邮箱"] = "%s@link.cuhk.edu.cn" % info["校园卡号"] if info["校园卡号"] != "卡号未知" else ""

            # 核对校园卡号是否重复
            if info["校园卡号"] in campus_idno_pool:
                idx_existed = campus_idno_pool[info["校园卡号"]]
                info_existed = infos[idx_existed]
                # 任意一人存在异常备注，则记录异常，不算重复，保留两条信息
                if len(info_existed["系统备注"]) != 0 or len(info["系统备注"]) != 0:
                    info["系统备注"] += "%s校园卡号重复且存在异常备注" % (
                        "" if len(info["系统备注"]) == 0 else "\n"
                    )
                    info_existed["系统备注"] += "%s校园卡号重复且存在异常备注" % (
                        "" if len(info_existed["系统备注"]) == 0 else "\n"
                    )
                # 两人姓名不同，则记录异常，不算重复，保留两条信息
                elif info["姓名"] != info_existed["姓名"]:
                    info["系统备注"] += "%s校园卡号重复但不同名 (%s | %s)" % (
                        "" if len(info["系统备注"]) == 0 else "\n",
                        info["姓名"],
                        info_existed["姓名"]
                    )
                    info_existed["系统备注"] += "%s校园卡号重复但不同名 (%s | %s)" % (
                        "" if len(info_existed["系统备注"]) == 0 else "\n",
                        info["姓名"],
                        info_existed["姓名"]
                    )
                # 算重复，合并其他信息
                else:
                    campus_idno_duplicated = True
                    for custom_title in custom_titles:
                        info_existed[custom_title] = "%s\n%s" % (
                            info_existed[custom_title],
                            info[custom_title]
                        )
                        infos[idx_existed] = info_existed

            # 储存不重复的信息
            if not campus_idno_duplicated:
                campus_idno_pool[info["校园卡号"]] = idx # 在对比池做记录
                infos[idx] = info # 储存

        # 重新排序
        infos_temp_1 = []
        infos_temp_2 = []
        for info in infos.values():
            if len(info["系统备注"]) != 0:
                infos_temp_1.append(info)
            else:
                infos_temp_2.append(info)
        infos_sort = []
        infos_sort += sorted(infos_temp_1, key=lambda x: x["校园卡号"])
        infos_sort += sorted(infos_temp_2, key=lambda x: x["校园卡号"])

    except Exception as e:
        print(traceback.format_exc())
        result["msgs"].append({
            "type": "error",
            "text": "读取文件失败<br>[读取文件] <= \"%s\"<br>[保存文件] => \"%s\"" % (
                old_wb_path, new_wb_path
            )
        })
        return result

    # 保存文件
    try:
        new_wb = Workbook()
        new_ws = new_wb.worksheets[0]

        new_aligntment = {
            "校园卡号": Alignment(
                vertical="center",
                horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "姓名": Alignment(
                vertical="center",
                horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "系统备注": Alignment(
                vertical="center",
                # horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "邮箱": Alignment(
                vertical="center",
                horizontal="center",
                # wrap_text=True # 自动换行
            ),
        }
        for custom_title in custom_titles:
            new_aligntment[custom_title] = Alignment(
                vertical="center",
                horizontal="center",
                # wrap_text=True # 自动换行
            )

        new_font = {
            "header": Font(
                bold=True
            )
        }

        new_ws.column_dimensions["A"].width = 20
        new_ws.column_dimensions["B"].width = 20
        new_ws.column_dimensions["C"].width = 60
        new_ws.column_dimensions["D"].width = 40

        for i in range(2, num_cols):
            new_ws.column_dimensions[get_column_letter(i+3)].width = 20

        new_titles = ["校园卡号", "姓名", "系统备注", "邮箱"] + custom_titles

        for col, key in enumerate(new_titles, start=1):
            cell = new_ws.cell(row=1, column=col)
            cell.value = key
            cell.alignment = new_aligntment[key]
            cell.font = new_font["header"]
            cell.number_format = numbers.FORMAT_TEXT

        for row, info in enumerate(infos_sort, start=2):
            for col, key in enumerate(new_titles, start=1):
                cell = new_ws.cell(row=row, column=col)
                cell.value = info[key]
                cell.alignment = new_aligntment[key]
                cell.number_format = numbers.FORMAT_TEXT

        new_wb.save(new_wb_path)

    except Exception as e:
        print(traceback.format_exc())
        result["msgs"].append({
            "type": "error",
            "text": "保存文件失败<br>[读取文件] <= \"%s\"<br>[保存文件] => \"%s\"" % (
                old_wb_path, new_wb_path
            )
        })
        return result

    result["msgs"].append({
        "type": "success",
        "text": "处理成功"
    })
    return result


if __name__ == "__main__":

    """
python -m utils.utils_tool_wenjuanxing2
    """

    old_wb_path = "./data/test/tool_utils-wenjuanxing-unprocessed.xlsx"
    new_wb_path = "./data/test/tool_utils-wenjuanxing-2.xlsx"

    result = preprocess_wenjuanxing_2(old_wb_path, new_wb_path)
    print(result["msgs"][-1]["type"])
    print("[读取文件] %s" % old_wb_path)
    print("[保存文件] %s" % new_wb_path)
