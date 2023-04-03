"""
Utils Course Record

# MongoDB 索引
mongo.coll_course_record.create_index(
    name="index", unique=True, keys=[("course_code", 1), ("campus_idno", 1)],
)
mongo.coll_course_record.create_index(
    name="index1", keys=[("course_code", 1)],
)
mongo.coll_course_record.create_index(
    name="index2", keys=[("campus_idno", 1)],
)

"""

import datetime
import traceback

from flask_login import current_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from .utils_file import file_utils
from .utils_mongo import mongo


class CourseRecordUtils(object):

    # ---
    # property
    # ---

    @property
    def new_doc(self):
        __new_doc = {
            "course_code":   "/",
            "campus_idno":   "/",

            "status":        "/",
            "activity_done": {},
            "authen":        "/",

            "note":          "备注",

            "createtime":    datetime.datetime.now(),
            "modifytime":    "/",
            "modifyuser":    "/",
        }
        return __new_doc

    @property
    def field_headers(self):
        __field_headers = {
            "course_code":   "课程代码", # unique with "campus_idno" & link to "course"
            "campus_idno":   "校园卡号", # unique with "course_code"

            "status":        "状态",     # update when modify "activity_record.count" or "activity.activity_type" or "course.activity_rules" or "course_record.authen"
            "activity_done": "活动进度", # update when modify "activity_record.count" or "activity.activity_type" or "course.activity_rules"
            "authen":        "认证",     # update manual only

            "note":          "备注",

            "createtime":    "创建时间",
            "modifytime":    "修改时间",
            "modifyuser":    "修改用户",
        }
        return __field_headers

    @property
    def field_options(self):
        __field_options = {
            "course_code": sorted(mongo.coll_course_record.distinct("course_code")),
            "status":      sorted(mongo.coll_course_record.distinct("status")),
        }
        return __field_options

    @property
    def field_limits(self):
        __field_limits = {
        }
        return __field_limits

    @property
    def field_fixeds(self):
        __field_fixeds = [
            "course_code",
            "campus_idno",

            "status",
            "activity_done",

            "createtime",
            "modifytime",
            "modifyuser",
        ]
        return __field_fixeds

    # ---
    # function.get
    # ---

    def get_record(self, course_code, campus_idno, update=False):
        return self.get_record_2(course_code, campus_idno, update)

    def get_record_1(self, course_code, campus_idno, update=False):
        result = {
            "msgs": [],
            "course_record": None,
        }

        # 校验字段 校园卡号
        if not mongo.coll_user_info.count_documents({"campus_idno": campus_idno}):
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>用户信息不存在 [ %s=%s ]" % (
                    self.field_headers["campus_idno"], campus_idno
                )
            })
            return result

        # 校验字段 课程代码
        if not mongo.coll_course_info.count_documents({"course_code": course_code}):
            msg = {
                "type": "error",
                "text": "校验失败<br>课程户信息不存在 [ %s=%s ]" % (
                    self.field_headers["course_code"], course_code
                )
            }
            result["msgs"].append(msg)
            return result

        # 新增 课程记录
        if not mongo.coll_course_record.count_documents({"course_code": course_code, "campus_idno": campus_idno}):
            __result = self.create_record(course_code=course_code, campus_idno=campus_idno)
            result["msgs"].extend(__result["msgs"])
            if __result["msgs"] and __result["msgs"][-1]["type"] != "success":
                return result
            __result = self.update_record(course_code=course_code, campus_idno=campus_idno)
            result["msgs"].extend(__result["msgs"])
            if __result["msgs"] and __result["msgs"][-1]["type"] == "error":
                return result

        # 更新 课程记录
        elif update:
            __result = self.update_record(course_code=course_code, campus_idno=campus_idno)
            result["msgs"].extend(__result["msgs"])
            if __result["msgs"] and __result["msgs"][-1]["type"] == "error":
                return result

        # 查询 课程记录
        course_records = list(mongo.coll_course_record.aggregate([
            {"$match": {"course_code": course_code, "campus_idno": campus_idno}},
            # 查询 人员信息
            {"$lookup": {
                "from": "coll_user_info", "as": "user_info",
                "localField": "campus_idno", "foreignField": "campus_idno"
            }},
            {"$unwind": "$user_info"},
            # 查询 课程信息
            {"$lookup": {
                "from": "coll_course_info", "as": "course_info",
                "localField": "course_code", "foreignField": "course_code"
            }},
            {"$unwind": "$course_info"},
        ]))
        if not course_records:
            return result
        course_record = course_records[0]

        # 查询 活动记录
        course_record["activity_records"] = list(mongo.coll_activity_record.aggregate([
            {"$match": {"campus_idno": campus_idno}},
            # 查询 活动信息
            {"$lookup": {
                "from": "coll_activity_info", "as": "activity_info",
                "localField": "activity_code", "foreignField": "activity_code"
            }},
            {"$unwind": "$activity_info"},
            {"$match": {"activity_info.course_code": course_code}},
            # 排序 活动信息
            {"$sort":{"activity_info.activity_code": 1}},
        ]))

        result["course_record"] = course_record
        return result

    def get_record_2(self, course_code, campus_idno, update=False):
        result = {
            "msgs": [],
            "course_record": None,
        }

        # 校验字段 校园卡号
        if not mongo.coll_user_info.count_documents({"campus_idno": campus_idno}):
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>用户信息不存在 [ %s=%s ]" % (
                    self.field_headers["campus_idno"], campus_idno
                )
            })
            return result

        # 校验字段 课程代码
        if not mongo.coll_course_info.count_documents({"course_code": course_code}):
            msg = {
                "type": "error",
                "text": "校验失败<br>课程户信息不存在 [ %s=%s ]" % (
                    self.field_headers["course_code"], course_code
                )
            }
            result["msgs"].append(msg)
            return result

        # 新增 课程记录
        if not mongo.coll_course_record.count_documents({"course_code": course_code, "campus_idno": campus_idno}):
            __result = self.create_record(course_code=course_code, campus_idno=campus_idno)
            result["msgs"].extend(__result["msgs"])
            if __result["msgs"] and __result["msgs"][-1]["type"] != "success":
                return result
            __result = self.update_record(course_code=course_code, campus_idno=campus_idno)
            result["msgs"].extend(__result["msgs"])
            if __result["msgs"] and __result["msgs"][-1]["type"] == "error":
                return result

        # 更新 课程记录
        elif update:
            __result = self.update_record(course_code=course_code, campus_idno=campus_idno)
            result["msgs"].extend(__result["msgs"])
            if __result["msgs"] and __result["msgs"][-1]["type"] == "error":
                return result

        # 查询 课程记录
        course_records = list(mongo.coll_course_record.aggregate([
            {"$match": {"course_code": course_code, "campus_idno": campus_idno}},
            # 查询 人员信息
            {"$lookup": {
                "from": "coll_user_info", "as": "user_info",
                "localField": "campus_idno", "foreignField": "campus_idno"
            }},
            {"$unwind": "$user_info"},
            # 查询 课程信息
            {"$lookup": {
                "from": "coll_course_info", "as": "course_info",
                "localField": "course_code", "foreignField": "course_code"
            }},
            {"$unwind": "$course_info"},
            # 查询 活动记录
            {"$lookup": {
                "from": "coll_activity_record", "as": "activity_records",
                "let": {
                    "campus_idno": "$campus_idno"
                },
                "pipeline": [
                    {"$match": {"$expr": {"$and": [
                        {"$eq": ["$campus_idno", "$$campus_idno"]}
                    ]}}},
                    # 查询 活动信息
                    {"$lookup": {
                        "from": "coll_activity_info", "as": "activity_info",
                        "localField": "activity_code", "foreignField": "activity_code"
                    }},
                    {"$unwind": "$activity_info"},
                    {"$match": {"activity_info.course_code": course_code}},
                    # 排序 活动信息
                    {"$sort":{"activity_info.activity_code": 1}},
                ]
            }}
        ]))
        if not course_records:
            return result
        course_record = course_records[0]

        result["course_record"] = course_record
        return result

    # ---
    # function.update
    # ---

    def update_record(self, course_code, campus_idno, modifyuser=None):
        result = {
            "msgs": []
        }

        # 校验 课程记录
        if not mongo.coll_course_record.count_documents({"course_code": course_code, "campus_idno": campus_idno}):
            result["msgs"].append({
                "type": "error",
                "text": "更新失败<br>课程记录不存在 [ %s=%s, %s=%s]" % (
                    self.field_headers["course_code"], course_code,
                    self.field_headers["campus_idno"], campus_idno,
                )
            })
            return result

        # 查询 课程信息
        course_info = mongo.coll_course_info.find_one({"course_code": course_code})

        # 查询 课程记录
        course_record = mongo.coll_course_record.find_one({"course_code": course_code, "campus_idno": campus_idno})

        # 查询 活动记录
        activity_records = list(mongo.coll_activity_record.aggregate([
            {"$match": {"campus_idno": campus_idno}},
            # 查询 活动信息
            {"$lookup": {
                "from": "coll_activity_info", "as": "activity_info",
                "localField": "activity_code", "foreignField": "activity_code"
            }},
            {"$unwind": "$activity_info"},
            {"$match": {"activity_info.course_code": course_code}},
        ]))

        # 获取 更新信息
        update_info = {}

        # 校验字段 特定字段 活动进度
        activity_done = {}
        for activity_record in activity_records:
            activity_type = activity_record["activity_info"]["activity_type"]
            activity_count = activity_record["count"]
            if activity_count:
                if activity_type not in activity_done:
                    activity_done[activity_type] = 0
                activity_done[activity_type] += activity_count
        if activity_done != course_record["activity_done"]:
            update_info["activity_done"] = activity_done

        # 校验字段 特定字段 状态
        if str(course_record["authen"]).upper() in ["TRUE", "T"]:
            status = "已认证"
        else:
            activity_rules = course_info["activity_rules"]
            is_doing = False
            is_done  = True
            for activity_rule in activity_rules:
                for activity_type in activity_rule:
                    rule_num = activity_rule[activity_type]
                    done_num = activity_done.get(activity_type, 0)
                    if done_num < rule_num:
                        is_done = False
                    if done_num > 0:
                        is_doing = True
                if is_done:
                    break
            if is_done:
                status = "已完成"
            elif is_doing:
                status = "进行中"
            else:
                status = "未开始"
        if status != course_record["status"]:
            update_info["status"] = status

        # 无需更新
        if len(update_info) == 0:
            msg = {
                "type": "warn",
                "text": "无需更新"
            }
            result["msgs"].append(msg)
            return result

        # 更新信息
        update_info["modifytime"] = datetime.datetime.now()
        update_info["modifyuser"] = modifyuser if modifyuser else current_user.idno
        __result = mongo.coll_course_record.update_one(
            {"course_code": course_code, "campus_idno": campus_idno},
            {"$set": update_info}
        )
        if __result.modified_count != 0:
            result["msgs"].append({
                "type": "success",
                "text": "更新成功<br>课程记录 [ 总数=1, 更新=%s, 错误=0 ]" % (
                    __result.modified_count
                )
            })
        else:
            result["msgs"].append({
                "type": "error",
                "text": "更新失败<br>课程记录 [ 总数=1, 更新=0, 错误=1 ]"
            })
        return result

    def update_record_batch(self):
        result = {
            "msgs": []
        }
        course_codes = mongo.coll_course_info.distinct("course_code")
        campus_idnos = mongo.coll_user_info.distinct("campus_idno", {"campus_role": "学生"})
        for course_code in course_codes:
            modified_count = 0
            unmodified_count = 0
            with tqdm(desc="[INFO] update_course_record [ course_code=%s ]" % course_code, total=len(campus_idnos)) as pbar:
                for campus_idno in campus_idnos:
                    if mongo.coll_course_record.count_documents({"course_code": course_code, "campus_idno": campus_idno}) == 0:
                        self.create_record(
                            course_code=course_code,
                            campus_idno=campus_idno
                        )
                    __result = self.update_record(
                        course_code=course_code,
                        campus_idno=campus_idno,
                        modifyuser="system"
                    )
                    if __result["msgs"][-1]["type"] == "success":
                        modified_count += 1
                    elif __result["msgs"][-1]["type"] == "warn":
                        unmodified_count += 1
                    else:
                        print(__result)
                    pbar.set_postfix(modified_count=modified_count, unmodified_count=unmodified_count)
                    pbar.update(1)
            if (modified_count + unmodified_count) == len(campus_idnos):
                result["msgs"].append({
                    "type": "success",
                    "text": "更新成功<br>%s => 课程记录 [ 总数=%s, 更新=%s, 无需更新=%s, 错误=%s ]" % (
                        course_code,
                        len(campus_idnos),
                        modified_count,
                        unmodified_count,
                        len(campus_idnos) - modified_count - unmodified_count
                    )
                })
            else:
                result["msgs"].append({
                    "type": "error",
                    "text": "更新失败<br>%s => 课程记录 [ 总数=%s, 更新=%s, 无需更新=%s, 错误=%s ]" % (
                        course_code,
                        len(campus_idnos),
                        modified_count,
                        unmodified_count,
                        len(campus_idnos) - modified_count - unmodified_count
                    )
                })
            print("[INFO] %s" % result["msgs"][-1]["text"].replace("<br>", "\n"))
        return result

    def update_record_by_file(self, update_file):
        result = {
            "msgs": [],
            "counts": {
                "total":    0,
                "update":   0,
                "insert":   0,
                "skip":     0,
                "delete":   0,
                "error":    0
            }
        }

        # 读取原始数据
        wb = load_workbook(update_file, data_only=True)
        ws = wb.worksheets[0]
        rows = []
        for row in ws.iter_rows():
            rows.append([cell.value for cell in row])

        row_header = rows[0]
        row_values = rows[1:]

        course_record_field_headers = self.field_headers
        course_record_field_limits  = self.field_limits
        course_record_field_fixeds  = self.field_fixeds

        course_record_header_fields = {header: field for field, header in course_record_field_headers.items()}

        # 校验字段 定位字段 活动代码
        if course_record_field_headers["course_code"] not in row_header:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少定位字段 [ %s ]" % course_record_field_headers["course_code"]
            })
            return result
        # 校验字段 定位字段 校园卡号
        if course_record_field_headers["campus_idno"] not in row_header:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少定位字段 [ %s ]" % course_record_field_headers["campus_idno"]
            })
            return result

        # 校验字段 无效字段
        invalid_headers = []
        for __header in row_header:
            if __header \
            and course_record_header_fields[__header] in course_record_field_headers \
            and (
                course_record_header_fields[__header] == "course_code" or \
                course_record_header_fields[__header] not in course_record_field_fixeds
            ):
                continue
            else:
                invalid_headers.append(__header)
                result["msgs"].append({
                    "type": "warn",
                    "text": "校验失败<br>存在无效字段 [ %s ]" % __header
                })

        # 校验字段 更新字段
        if len(row_header) - len(invalid_headers) <= 1:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少更新字段"
            })
            return result

        with tqdm(desc="[INFO] update_course_record_by_file", total=len(row_values)) as pbar:

            for row_value in row_values:
                result["counts"]["total"] += 1
                pbar.update(1)

                # 获取更新信息
                delete        = False # 删除
                course_code = None  # 活动代码
                campus_idno   = None  # 校园卡号

                update_info = {}
                default_doc = self.new_doc
                for idx, value in enumerate(row_value):
                    header = row_header[idx]
                    # 转换格式
                    if isinstance(value, (int, float)):
                        value = str(value)
                    # 提取字段 删除
                    if header == "删除":
                        if str(value).upper() in ["TRUE", "T"]:
                            delete = True
                            continue
                    # 校验字段 未知字段 跳过
                    if header not in course_record_header_fields:
                        continue
                    # 转义字段
                    field = course_record_header_fields[header]
                    # 提取字段 活动代码
                    if field == "course_code":
                        course_code = value
                        continue
                    # 提取字段 活动代码
                    if field == "campus_idno":
                        campus_idno = value
                        continue
                    # 校验字段 固定字段 跳过
                    if field in course_record_field_fixeds:
                        continue
                    # 校验字段 所有字段 默认值
                    if not value:
                        value = default_doc[field]
                    # 校验字段 限制字段
                    if field in course_record_field_limits and value not in course_record_field_limits[field]:
                        result["msgs"].append({
                            "type": "error",
                            "text": "校验失败<br>值不可选 [ %s=%s ]" % (
                                course_record_field_headers[field], value
                            )
                        })
                        continue
                    # 记录字段
                    update_info[field] = value

                # 校验字段 定位字段 活动代码
                if mongo.coll_course_info.count_documents({"course_code": course_code}) == 0:
                    result["counts"]["error"] += 1
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值不存在 [ %s=%s ]" % (
                            course_record_field_headers["course_code"], course_code
                        )
                    })
                    continue
                # 校验字段 定位字段 校园卡号
                if mongo.coll_user_info.count_documents({"campus_idno": campus_idno}) == 0:
                    result["counts"]["error"] += 1
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值不存在 [ %s=%s ]" % (
                            course_record_field_headers["campus_idno"], campus_idno
                        )
                    })
                    continue

                course_record = mongo.coll_course_record.find_one({"course_code": course_code, "campus_idno": campus_idno})

                # 删除
                if delete:

                    if course_record:
                        delete_result = mongo.coll_course_record.delete_one({"course_code": course_code, "campus_idno": campus_idno})
                        result["counts"]["delete"] += 1
                        result["msgs"].append({
                            "type": "warn",
                            "text": "删除成功<br>[ %s=%s, %s=%s ] 活动记录: %s" % (
                                course_record_field_headers["course_code"], course_code,
                                course_record_field_headers["campus_idno"],   campus_idno,
                                delete_result.deleted_count,
                            )
                        })

                    else:
                        result["counts"]["skip"] += 1
                        result["msgs"].append({
                            "type": "warn",
                            "text": "删除失败<br>[ %s=%s, %s=%s ] 活动记录不存在" % (
                                course_record_field_headers["course_code"], course_code,
                                course_record_field_headers["campus_idno"],   campus_idno,
                            )
                        })

                else:

                    # 更新
                    if course_record:

                        # 确认更新
                        update_info = {field: value for field, value in update_info.items() if value != course_record[field]}

                        # 无需更新
                        if not update_info:
                            result["counts"]["skip"] += 1
                            continue

                        # 更新信息
                        update_info["modifytime"] = datetime.datetime.now()
                        update_info["modifyuser"] = current_user.idno
                        update_result = mongo.coll_course_record.update_one(
                            {"course_code": course_code, "campus_idno": campus_idno},
                            {"$set": update_info}
                        )
                        if update_result.modified_count == 1:
                            result["counts"]["update"] += 1

                            # 同步更新
                            if "authen" in update_info:
                                # 同步更新 认证 => 课程记录
                                __result = course_record_utils.update_record(
                                    course_code=course_code,
                                    campus_idno=campus_idno
                                )
                                if __result["msgs"][-1]["type"] == "error":
                                    print("[ERROR] %s" % __result["msgs"][-1]["text"])

                        else:
                            result["counts"]["error"] += 1
                            result["msgs"].append({
                                "type": "error",
                                "text": "更新失败<br>[ %s=%s, %s=%s ]" % (
                                    course_record_field_headers["course_code"], course_code,
                                    course_record_field_headers["campus_idno"],   campus_idno,
                                )
                            })
                            print("[ERROR] update_course_record_by_file [ course_code=%s, modified_count: %s ]\n%s" % (
                                course_code,
                                update_result.modified_count,
                                update_info
                            ))

                    # 新增
                    else:

                        # 获取 新增信息
                        insert_info = self.new_doc
                        insert_info["course_code"] = course_code
                        insert_info["campus_idno"] = campus_idno

                        # 新增信息
                        insert_info.update(update_info)
                        insert_info["modifytime"] = datetime.datetime.now()
                        insert_info["modifyuser"] = current_user.idno
                        mongo.coll_course_record.insert_one(insert_info)
                        result["counts"]["insert"] += 1

                        # 同步更新 课程记录
                        __result = course_record_utils.update_record(
                            course_code=course_code,
                            campus_idno=campus_idno
                        )
                        if __result["msgs"][-1]["type"] == "error":
                            print("[ERROR] %s" % __result["msgs"][-1]["text"])

        result["msgs"] = [{
            "type": "success",
            "text": "更新成功<br>活动记录 [ 总数=%s, 更新=%s, 新增=%s, 跳过=%s, 删除=%s, 错误=%s]" % (
                result["counts"]["total"],
                result["counts"]["update"],
                result["counts"]["insert"],
                result["counts"]["skip"],
                result["counts"]["delete"],
                result["counts"]["error"],
            )
        }] + result["msgs"]

        print("[INFO] total: %s, update: %s, insert: %s, skip: %s, delete: %s, error: %s" % (
            result["counts"]["total"],
            result["counts"]["update"],
            result["counts"]["insert"],
            result["counts"]["skip"],
            result["counts"]["delete"],
            result["counts"]["error"],
        ))

        return result

    # ---
    # function.create
    # ---

    def create_record(self, course_code, campus_idno):
        result = {
            "msgs": [],
        }
        # 获取 新增信息
        insert_record = self.new_doc
        # 新增信息
        if course_code:
            insert_record["course_code"] = course_code
        if campus_idno:
            insert_record["campus_idno"] = campus_idno
        insert_record["authen"] = "False"
        insert_record["status"] = "未开始"
        mongo.coll_course_record.insert_one(insert_record)
        result["msgs"].append({
            "type": "success",
            "text": "新增成功<br>课程记录 [ %s=%s, %s=%s ]" % (
                self.field_headers["course_code"], course_code,
                self.field_headers["campus_idno"], campus_idno
            )
        })
        return result

    # ---
    # function.save
    # ---

    def save_records(self, course_records):
        result = {
            "msgs": [],
            "file_info": None,
        }

        file_dir = "temp_dir"
        date_info = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = "course_record_utils-save_records-%s.xlsx" % date_info
        save_name = "课程记录-%s.xlsx" % date_info
        file_path = file_utils.get_path(file_dir=file_dir, file_name=file_name)

        try:
            wb = Workbook()
            ws = wb.worksheets[0]

            header_fields = {header: field for field, header in self.field_headers.items()}

            alignment = Alignment(
                vertical="center",
                horizontal="center"
            )

            aligntments = {header: alignment for header in header_fields}

            fonts = {
                "header": Font(
                    bold=True
                )
            }

            for col, _ in enumerate(header_fields):
                ws.column_dimensions[get_column_letter(col+1)].width = 20

            for col, header in enumerate(header_fields, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.alignment = aligntments[header]
                cell.font = fonts["header"]
                cell.number_format = numbers.FORMAT_TEXT

            for row, course_record in enumerate(course_records, start=2):
                for col, header in enumerate(header_fields, start=1):
                    field = header_fields[header]
                    cell = ws.cell(row=row, column=col)
                    cell.value = str(course_record[field]).strip().replace("'", '"')
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.alignment = aligntments[header]

            wb.save(file_path)

        except Exception as e:
            print(traceback.format_exc())
            result["msgs"].append({
                "type": "error",
                "text": "保存文件失败<br>[ file_dir=%s, file_name=%s ]" % (file_dir, file_name)
            })
            return result

        result["file_info"] = {
            "file_dir": file_dir,
            "file_name": file_name,
            "save_name": save_name
        }
        return result


    def save_records_with_detail(self, course_records):
        from .utils_user import user_utils

        result = {
            "msgs": [],
            "file_info": None,
        }

        file_dir = "temp_dir"
        date_info = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = "course_record_utils-save_records_with_detail-%s.xlsx" % date_info
        save_name = "课程记录[详细]-%s.xlsx" % date_info
        file_path = file_utils.get_path(file_dir=file_dir, file_name=file_name)

        try:
            wb = Workbook()
            ws = wb.worksheets[0]

            user_info_field_headers = user_utils.field_headers
            for __field in [
                "idno",
                "sex",
                "bankacc",
                "phoneno",

                "createtime",
                "modifytime",
                "modifyuser",

                "modifykeys",
            ]:
                user_info_field_headers.pop(__field)
            user_header_fields = {header: field for field, header in user_info_field_headers.items()}

            course_record_field_headers = self.field_headers
            for __field in [
                "campus_idno",

                "createtime",
                "modifytime",
                "modifyuser",
            ]:
                course_record_field_headers.pop(__field)
            course_record_header_fields = {header: field for field, header in course_record_field_headers.items()}

            alignment = Alignment(
                vertical="center",
                horizontal="center"
            )

            aligntments = {}
            aligntments.update({header: alignment for header in user_header_fields})
            aligntments.update({header: alignment for header in course_record_header_fields})

            fonts = {
                "header": Font(
                    bold=True
                )
            }

            # 列宽
            for col, _ in enumerate(user_header_fields):
                ws.column_dimensions[get_column_letter(col+1)].width = 20
            for col, _ in enumerate(course_record_header_fields, start=len(user_header_fields)):
                ws.column_dimensions[get_column_letter(col+1)].width = 20

            # 标题
            for col, header in enumerate(user_header_fields, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.alignment = aligntments[header]
                cell.font = fonts["header"]
                cell.number_format = numbers.FORMAT_TEXT
            for col, header in enumerate(course_record_header_fields, start=len(user_header_fields)+1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.alignment = aligntments[header]
                cell.font = fonts["header"]
                cell.number_format = numbers.FORMAT_TEXT

            # 内容
            for row, course_record in enumerate(course_records, start=2):
                for col, header in enumerate(user_header_fields, start=1):
                    field = user_header_fields[header]
                    cell = ws.cell(row=row, column=col)
                    cell.value = str(course_record["user_info"][field]).strip().replace("'", '"')
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.alignment = aligntments[header]
                for col, header in enumerate(course_record_header_fields, start=len(user_header_fields)+1):
                    field = course_record_header_fields[header]
                    cell = ws.cell(row=row, column=col)
                    cell.value = str(course_record[field]).strip().replace("'", '"')
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.alignment = aligntments[header]

            wb.save(file_path)

        except Exception as e:
            print(traceback.format_exc())
            result["msgs"].append({
                "type": "error",
                "text": "保存文件失败<br>[ file_dir=%s, file_name=%s ]" % (file_dir, file_name)
            })
            return result

        result["file_info"] = {
            "file_dir": file_dir,
            "file_name": file_name,
            "save_name": save_name
        }
        return result


course_record_utils = CourseRecordUtils()


if __name__ == "__main__":
    """
python -m app.utils.utils_course_record
    """

    course_record_utils.update_record_batch()
