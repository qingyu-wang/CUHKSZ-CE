"""
openpyxl number_format
https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
"""

import datetime
import traceback

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, numbers

from .utils_mongo import mongo


def get_worksheet_value_by_rows(ws):
    values = []
    for row in ws.iter_rows():
        values.append([cell.value for cell in row])
    return values


def preprocess_tencent(old_wb_path, new_wb_path):
    result = {
        "msgs": [],
    }

    # 读取文件
    try:
        # 读取原始数据
        old_wb = load_workbook(old_wb_path)
        old_ws = old_wb.worksheets[0]
        old_raw_values = get_worksheet_value_by_rows(old_ws)
        old_titles = {title: i for i, title in enumerate(old_raw_values[0])}
        old_values = old_raw_values[1:]

        # 清洗数据
        infos = {}
        for idx, value in enumerate(old_values):
            text = value[old_titles["用户昵称（入会昵称）"]]
            # 解析学号
            temps = []

            if len(text) >= 9 and text[:9].strip().isdigit():
                temps.append(text[:9].strip())

            split_1 = text.split("(")
            if len(split_1) >= 1 and len(split_1[0]) >= 9  and split_1[0][-9:].isdigit():
                temps.append(split_1[0][-9:])
            if len(split_1) >= 2 and len(split_1[1]) >= 9  and split_1[1][:9].isdigit():
                temps.append(split_1[1][:9])
            if len(split_1) >= 2 and len(split_1[1]) >= 14 and split_1[1][5:14].isdigit():
                temps.append(split_1[1][5:14])
            if len(split_1) >= 3 and len(split_1[2]) >= 9  and split_1[2][:9].isdigit():
                temps.append(split_1[2][:9])
            if len(split_1) >= 4 and len(split_1[3]) >= 9  and split_1[3][:9].isdigit():
                temps.append(split_1[3][:9])

            split_2 = text.split(")")
            if len(split_2) >= 1 and len(split_2[0]) >= 9  and split_2[0][-9:].isdigit():
                temps.append(split_2[0][-9:])
            if len(split_2) >= 2 and len(split_2[1]) >= 9  and split_2[1][-9:].isdigit():
                temps.append(split_2[1][-9:])

            # 确定学号
            if temps:
                campus_idno = temps[0]
                for temp in temps[1:]:
                    if temp != campus_idno:
                        campus_idno = "[无法解析] %s" % text
                        break
                if mongo.coll_user_info.count_documents({"campus_role": "学生", "campus_idno": campus_idno}) != 1:
                    campus_idno = "[ERROR: 校园卡号不存在] %s" % text
            else:
                campus_idno = "[ERROR: 无法解析] %s" % text

            # 签到记录
            time_signin = value[old_titles["入会时间"]]
            time_signin = datetime.datetime.strptime(time_signin, "%Y-%m-%d %H:%M:%S")
            time_signout = value[old_titles["退会时间"]]
            if time_signout == "--":
                time_signout = time_signin
            else:
                time_signout = datetime.datetime.strptime(time_signout, "%Y-%m-%d %H:%M:%S")
            # 持续时间
            duration = time_signout - time_signin
            # 更新信息
            if campus_idno not in infos:
                infos[campus_idno] = {
                    "校园卡号": campus_idno,
                    "持续时间": duration,
                    "签到记录": [time_signin],
                    "签退记录": [time_signout]
                }
            else:
                infos[campus_idno]["持续时间"] += duration
                infos[campus_idno]["签到记录"].append(time_signin)
                infos[campus_idno]["签退记录"].append(time_signout)

        # 重新排序
        infos_temp_1 = []
        infos_temp_2 = []
        for info in infos.values():
            if "ERROR" in info["校园卡号"]:
                infos_temp_1.append(info)
            else:
                infos_temp_2.append(info)
        infos_sort = []
        infos_sort += sorted(infos_temp_1, key=lambda x: x["持续时间"], reverse=True)
        infos_sort += sorted(infos_temp_2, key=lambda x: x["持续时间"], reverse=True)

    except Exception as e:
        print(traceback.format_exc())
        result["msgs"].append({
            "type": "error",
            "text": "读取文件失败<br>[读取文件] <= \"%s\"<br>[保存文件] => \"%s\"" % (
                old_wb_path, new_wb_path
            )
        })
        return result

    # 保存文件
    try:
        new_wb = Workbook()
        new_ws = new_wb.worksheets[0]

        headers = ["校园卡号", "持续时间 (秒)", "持续时间", "签到记录", "签退记录"]

        new_aligntment = {
            "校园卡号": Alignment(
                vertical="center",
                # horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "持续时间 (秒)": Alignment(
                vertical="center",
                horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "持续时间": Alignment(
                vertical="center",
                horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "签到记录": Alignment(
                vertical="center",
                # horizontal="center",
                # wrap_text=True # 自动换行
            ),
            "签退记录": Alignment(
                vertical="center",
                # horizontal="center",
                # wrap_text=True # 自动换行
            ),
        }
        new_font = {
            "header": Font(
                bold=True
            )
        }

        new_ws.column_dimensions["A"].width = 20
        new_ws.column_dimensions["B"].width = 20
        new_ws.column_dimensions["C"].width = 20
        new_ws.column_dimensions["D"].width = 20
        new_ws.column_dimensions["E"].width = 20

        for col, key in enumerate(headers, start=1):
            cell = new_ws.cell(row=1, column=col)
            cell.value = key
            cell.alignment = new_aligntment[key]
            cell.font = new_font["header"]
            cell.number_format = numbers.FORMAT_TEXT

        for row, info in enumerate(infos_sort, start=2):
            for col, key in enumerate(headers, start=1):
                cell = new_ws.cell(row=row, column=col)
                if key == "签到记录":
                    cell.value = "\n".join([j.strftime("%Y-%m-%d %H:%M:%S") for j in info[key]])
                    cell.number_format = numbers.FORMAT_TEXT
                elif key == "签退记录":
                    cell.value = "\n".join([j.strftime("%Y-%m-%d %H:%M:%S") for j in info[key]])
                    cell.number_format = numbers.FORMAT_TEXT
                elif key == "持续时间 (秒)":
                    cell.value = info["持续时间"].total_seconds()
                    cell.number_format = numbers.FORMAT_NUMBER
                elif key == "持续时间":
                    cell.value = str(info[key])
                    cell.number_format = numbers.FORMAT_TEXT
                else:
                    cell.value = info[key]
                    cell.number_format = numbers.FORMAT_TEXT
                cell.alignment = new_aligntment[key]

        new_wb.save(new_wb_path)

    except Exception as e:
        print(traceback.format_exc())
        result["msgs"].append({
            "type": "error",
            "text": "保存文件失败<br>[读取文件] <= \"%s\"<br>[保存文件] => \"%s\"" % (
                old_wb_path, new_wb_path
            )
        })
        return result

    result["msgs"].append({
        "type": "success",
        "text": "处理成功"
    })
    return result


if __name__ == "__main__":

    """
python -m utils.utils_tool_tencent
    """

    old_wb_path = "./data/test/tool_utils-tencent-unprocessed.xlsx"
    new_wb_path = "./data/test/tool_utils-tencent.xlsx"

    result = preprocess_tencent(old_wb_path, new_wb_path)
    print(result["msgs"][-1]["type"])
    print("[读取文件] %s" % old_wb_path)
    print("[保存文件] %s" % new_wb_path)
