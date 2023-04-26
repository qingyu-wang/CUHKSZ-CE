"""
Utils User

# MongoDB 索引
mongo.coll_user_info.create_index(
    name="index", unique=True, keys=[("campus_idno", 1)]
)

"""

import datetime
import traceback

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from flask_login import current_user

from .utils_file import file_utils
from .utils_mongo import mongo


class UserUtils(object):

    # ---
    # property
    # ---

    @property
    def new_doc(self):
        __new_doc = {
            "campus_idno":   "/",
            "campus_role":   "/",

            "name":          "/",
            "campus_type":   "/",
            "campus_dept":   "/",
            "campus_addr":   "/",
            "campus_year":   "/",
            "campus_grade":  "/",
            "campus_source": "/",
            "campus_status": "/",

            "idno":          "/",
            "sex":           "/",
            "bankacc":       "/",
            "phoneno":       "/",

            "createtime":    datetime.datetime.now(),
            "modifytime":    "/",
            "modifyuser":    "/",
            "modifykeys":    [],
        }
        return __new_doc

    @property
    def field_headers(self):
        __field_headers = {
            "campus_idno":   "校园卡号", # unique & link to "activity_record", "course_record"
            "campus_role":   "角色",

            "name":          "姓名",
            "campus_type":   "类型",
            "campus_dept":   "部门",
            "campus_addr":   "书院",
            "campus_year":   "入学年份",
            "campus_grade":  "年级",
            "campus_source": "生源",
            "campus_status": "修读状态",

            "idno":          "身份证号",
            "sex":           "性别",
            "bankacc":       "银行账号",
            "phoneno":       "电话号码",

            "createtime":    "创建时间",
            "modifytime":    "修改时间",
            "modifyuser":    "修改用户",

            "modifykeys":    "锁定字段"
        }
        return __field_headers

    @property
    def field_options(self):
        __field_options = {
            "campus_role":   sorted(mongo.coll_user_info.distinct("campus_role")),

            "campus_type":   sorted(mongo.coll_user_info.distinct("campus_type")),
            "campus_dept":   sorted(mongo.coll_user_info.distinct("campus_dept")),
            "campus_addr":   sorted(mongo.coll_user_info.distinct("campus_addr")),
            "campus_year":   sorted(mongo.coll_user_info.distinct("campus_year")),
            "campus_grade":  sorted(mongo.coll_user_info.distinct("campus_grade")),
            "campus_source": sorted(mongo.coll_user_info.distinct("campus_source")),
            "campus_status": sorted(mongo.coll_user_info.distinct("campus_status")),

            "sex":           sorted(mongo.coll_user_info.distinct("sex")),
        }
        return __field_options

    @property
    def field_limits(self):
        __field_limits = {
            "campus_role": ["学生", "教工"],
        }
        return __field_limits

    @property
    def field_fixeds(self):
        __field_fixeds = [
            "createtime",
            "modifytime",
            "modifyuser",
        ]
        return __field_fixeds

    # ---
    # function.get
    # ---

    def get_info_by_campus_idno(self, campus_idno):
        result = {
            "msgs": [],
            "user_info": None,
        }
        user_info = mongo.coll_user_info.find_one({"campus_idno": campus_idno})
        result["user_info"] = user_info
        return result

    def get_info_by_campus_idno_with_detail(self, campus_idno):
        from .utils_course_record import course_record_utils

        result = {
            "msgs": [],
            "user_info": None,
        }
        # 查询用户信息
        user_info = mongo.coll_user_info.find_one({"campus_idno": campus_idno})

        if not user_info:
            pass
            # msg = {
            #     "type": "warn",
            #     "text": "查询失败: 用户不存在<br>[ %s=%s ]" % (
            #         user_utils.field_headers["campus_idno"], campus_idno,
            #     )
            # }
            # result["msgs"].append(msg)

        # 角色 学生
        elif user_info["campus_role"] == "学生":

            for course_code in mongo.coll_course_info.distinct("course_code"):
                if mongo.coll_course_record.count_documents({"course_code": course_code, "campus_idno": campus_idno}) == 0:
                    # 新增课程记录
                    __result = course_record_utils.create_record(
                        course_code=course_code, 
                        campus_idno=campus_idno
                    )
                    result["msgs"].extend(__result["msgs"])
                    if __result["msgs"] and __result["msgs"][-1]["type"] != "success":
                        return result
                    # 刷新课程记录
                    __result = course_record_utils.update_record(
                        course_code=course_code,
                        campus_idno=campus_idno
                    )
                    result["msgs"].extend(__result["msgs"])
                    if __result["msgs"] and __result["msgs"][-1]["type"] != "success":
                        return result

            # 查询信息
            user_infos = list(mongo.coll_user_info.aggregate([
                # 查询 校园卡号
                {"$match": {"campus_idno": campus_idno}},
                # 查询 课程记录
                {"$lookup": {
                    "from": "coll_course_record", "as": "course_records",
                    "let": {"campus_idno": "$campus_idno"},
                    "pipeline": [
                        {"$match": {"$expr": {"$and": [
                            {"$eq": ["$campus_idno", campus_idno]}
                        ]}}},
                        # 查询 课程信息
                        {"$lookup": {
                            "from": "coll_course_info", "as": "course_info",
                            "localField": "course_code","foreignField": "course_code"
                        }},
                        {"$unwind": "$course_info"},
                        # 排序
                        {"$sort": {"course_code": 1}}
                    ],
                }},
                # 查询 活动记录
                {"$lookup": {"from": "coll_activity_record", "as": "activity_records",
                    "let": {"campus_idno": "$campus_idno"},
                    "pipeline": [
                        {"$match": {"$expr": {"$and": [
                            {"$eq": ["$campus_idno", campus_idno]}
                        ]}}},
                        # 查询 活动信息
                        {"$lookup": {
                            "from": "coll_activity_info", "as": "activity_info",
                            "localField": "activity_code","foreignField": "activity_code"
                        }},
                        {"$unwind": "$activity_info"},
                        # 查询 课程信息
                        {"$lookup": {
                            "from": "coll_course_info", "as": "course_info",
                            "localField": "activity_info.course_code", "foreignField": "course_code"
                        }},
                        {"$unwind": "$course_info"},
                        # 排序
                        {"$sort": {"activity_code": 1}}
                    ]
                }},
            ]))

            if not user_infos:
                result["user_info"] = user_info
                return result

            user_info = user_infos[0]

            # 后处理
            __course_records = {}
            for course_record in user_info["course_records"]:
                course_record["activity_records"] = []
                __course_records[course_record["course_code"]] = course_record

            __activity_records = user_info.pop("activity_records")

            for __activity_record in __activity_records:
                __course_code = __activity_record["course_info"]["course_code"]
                __course_record = __course_records[__course_code]
                __course_record["activity_records"].append(__activity_record)

            user_info["course_records"] = list(__course_records.values())
            result["user_info"] = user_info

        # 角色 教工
        else:
            result["user_info"] = user_info

        return result

    def get_infos_by_name(self, name):
        result = {
            "msgs": [],
            "user_infos": [],
        }
        user_infos = list(mongo.coll_user_info.find({"name": name}))
        result["user_infos"] = user_infos
        return result

    def get_infos(self, field_filters):
        result = {
            "msgs": [],
            "user_infos": [],
        }
        user_infos = list(mongo.coll_user_info.find(field_filters).sort([("campus_idno", 1)]))
        result["user_infos"] = user_infos
        return result

    # ---
    # function.update
    # ---

    def update_info(self, new_user_info, old_campus_idno):
        result = {
            "msgs": []
        }
        old_user_info = mongo.coll_user_info.find_one({"campus_idno": old_campus_idno})

        # 获取 更新信息
        update_info = {}
        update_info["modifykeys"] = old_user_info["modifykeys"]

        user_info_field_headers = self.field_headers
        user_info_field_limits  = self.field_limits
        user_info_field_fixeds  = self.field_fixeds

        default_doc = self.new_doc

        for field in new_user_info:
            new_value = new_user_info[field]
            old_value = old_user_info[field]
            # 校验字段 固定字段 跳过
            if field in user_info_field_fixeds:
                continue
            # 校验字段 所有字段 默认值
            if not new_value:
                new_value = default_doc[field]
            # 校验字段 特定字段 校园卡号
            if field == "campus_idno" and new_value != old_value:
                user_num = mongo.coll_user_info.count_documents({"campus_idno": new_value})
                if user_num != 0:
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值已存在 [ %s=%s ]" % (user_info_field_headers[field], new_value)
                    })
                    return result
            # 校验字段 限制字段
            if field in user_info_field_limits and new_value not in user_info_field_limits[field]:
                result["msgs"].append({
                    "type": "error",
                    "text": "校验失败<br>值不可选 [ %s=%s ]" % (user_info_field_headers[field], new_value)
                })
                return result
            # 更新字段
            if new_value != old_value:
                update_info[field] = new_value
                # 手动更新后，锁定字段 (无法自动更新)
                if field not in update_info["modifykeys"]:
                    update_info["modifykeys"].append(field)

        update_info["modifykeys"] = list(set(update_info["modifykeys"]))
        update_info["modifykeys"].sort()

        # 无需更新
        if len(update_info) == 1: # 只包含 modifykeys
            msg = {
                "type": "warn",
                "text": "无需更新"
            }
            result["msgs"].append(msg)
            return result

        # 更新信息
        update_info["modifytime"] = datetime.datetime.now()
        update_info["modifyuser"] = current_user.idno
        update_result = mongo.coll_user_info.update_one(
            {"campus_idno": old_campus_idno},
            {"$set": update_info}
        )
        if update_result.modified_count != 0:
            result["msgs"].append({
                "type": "success",
                "text": "更新成功<br>人员信息 [ 总数=1, 更新=%s, 错误=0 ]" % update_result.modified_count
            })
        else:
            result["msgs"].append({
                "type": "error",
                "text": "更新失败<br>人员信息 [ 总数=1, 更新=0, 错误=1 ]"
            })
        return result

    def update_info_by_file(self, update_file):
        result = {
            "msgs": [],
            "counts": {
                "total":    0,
                "update":   0,
                "insert":   0,
                "skip":     0,
                "delete":   0,
                "error":    0
            }
        }

        filed_value_map = {
            "campus_dept": {
                "HSS":  "人文社科学院 / HSS",
                "SHSS": "人文社科学院 / HSS",
                "LHS":  "生命健康学院 / LHS",
                "MED":  "医学院 / MED",
                "MUS":  "音乐学院 / MUS",
                "SDS":  "数据科学学院 / SDS",
                "SME":  "经管学院 / SME",
                "SSE":  "理工学院 / SSE",
            },
            "campus_addr": {
                "Shaw College":         "逸夫书院 / SHAW",
                "Diligentia College":   "学勤书院 / DILIGENTIA",
                "Muse College":         "思廷书院 / MUSE",
                "Harmonia College":     "祥波书院 / HARMONIA",
                "Ling College":         "道扬书院 / LING",
            },
            "campus_grade": {
                "First Year":   "1年级",
                "Second Year":  "2年级",
                "Third Year":   "3年级",
                "Fourth Year":  "4年级",
            },
            "campus_source": {
                "631":                      "1 国内 - 631",
                "非631":                    "2 国内 - 高考",
                "国内本科生-保送":          "3 国内 - 保送",
                "艺术类本科批":             "4 国内 - 艺术",
                "Special Admission of UG":  "5 国内 - 海本",
                "国际生":                   "6 国际",
            },
            "campus_status": {
                "Active in Program":    "1 修读",
                "Completed Program":    "2 毕业",
                "Leave of Absence":     "3 休学",
                "Cancelled":            "4 退学",
                "Discontinued":         "4 退学",
            }
        }

        # 读取原始数据
        wb = load_workbook(update_file, data_only=True)
        ws = wb.worksheets[0]
        rows = []
        for row in ws.iter_rows():
            rows.append([cell.value for cell in row])

        row_header = rows[0]
        row_values = rows[1:]

        user_info_field_headers = self.field_headers
        user_info_field_limits  = self.field_limits
        user_info_field_fixeds  = self.field_fixeds

        user_header_fields = {header: field for field, header in user_info_field_headers.items()}

        # 校验字段 定位字段 校园卡号
        if user_info_field_headers["campus_idno"] not in row_header:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少定位字段 [ %s ]" % user_info_field_headers["campus_idno"]
            })
            return result

        # 校验字段 无效字段
        invalid_headers = []
        for __header in row_header:
            if __header == "删除":
                continue
            elif __header \
            and user_header_fields[__header] in user_info_field_headers \
            and (
                user_header_fields[__header] == "campus_idno" or \
                user_header_fields[__header] not in user_info_field_fixeds
            ):
                continue
            else:
                invalid_headers.append(__header)
                result["msgs"].append({
                    "type": "warn",
                    "text": "校验失败<br>存在无效字段 [ %s ]" % __header
                })

        # 校验字段 缺少有效字段
        if len(row_header) - len(invalid_headers) <= 1:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少有效字段"
            })
            return result

        with tqdm(desc="[INFO] update_user_info_by_file", total=len(row_values)) as pbar:

            for row_value in row_values:
                result["counts"]["total"] += 1
                pbar.update(1)

                # 获取更新信息
                delete      = False # 删除
                campus_idno = None  # 校园卡号
                modifykeys  = None  # 锁定字段

                update_info = {}
                default_doc = self.new_doc
                for idx, value in enumerate(row_value):
                    header = row_header[idx]
                    # 提取字段 删除
                    if header == "删除":
                        if str(value).upper() in ["TRUE", "T"]:
                            delete = True
                            continue
                    # 校验字段 未知字段 跳过
                    if header not in user_header_fields:
                        continue
                    # 转义字段
                    field = user_header_fields[header]
                    # 提取字段 校园卡号
                    if field == "campus_idno":
                        campus_idno = value
                        continue
                    # 提取字段 锁定字段
                    if field == "modifykeys":
                        if value:
                            modifykeys = [user_header_fields[key.strip()] for key in value.split(",") if key.strip() in user_header_fields]
                        else:
                            modifykeys = []
                        continue
                    # 校验字段 固定字段 跳过
                    if field in user_info_field_fixeds:
                        continue
                    # 校验字段 所有字段 默认值
                    if not value:
                        value = default_doc[field]
                    # 校验字段 预处理字段
                    if field in filed_value_map:
                        if value in filed_value_map[field]:
                            value = filed_value_map[field][value]
                    # 校验字段 限制字段
                    if field in user_info_field_limits and value not in user_info_field_limits[field]:
                        result["msgs"].append({
                            "type": "error",
                            "text": "校验失败<br>值不可选 [ %s=%s ]" % (user_info_field_headers[field], value)
                        })
                        continue
                    # 记录字段
                    update_info[field] = value

                # 校验字段 定位字段 校园卡号
                if not campus_idno:
                    result["counts"]["error"] += 1
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值不能为空 [ %s ]" % user_info_field_headers["campus_idno"]
                    })
                    continue

                user_info = mongo.coll_user_info.find_one({"campus_idno": campus_idno})

                # 删除
                if delete:

                    if user_info:
                        # 有"活动记录"不能删除"人员信息"
                        if mongo.coll_activity_record.count_documents({"campus_idno": campus_idno}) != 0:
                            result["counts"]["error"] += 1
                            result["msgs"].append({
                                "type": "error",
                                "text": "删除失败<br>[ %s=%s ] 存在活动记录" % (
                                    user_info_field_headers["campus_idno"], campus_idno
                                )
                            })

                        else:
                            delete_result_1 = mongo.coll_user_info.delete_one({"campus_idno": campus_idno})
                            delete_result_2 = mongo.coll_course_record.delete_many({"campus_idno": campus_idno})
                            result["counts"]["delete"] += 1
                            result["msgs"].append({
                                "type": "success",
                                "text": "删除成功<br>[ %s=%s ] 人员信息: %s, 课程记录: %s" % (
                                    user_info_field_headers["campus_idno"], campus_idno,
                                    delete_result_1.deleted_count,
                                    delete_result_2.deleted_count,
                                )
                            })

                    else:
                        result["counts"]["skip"] += 1
                        result["msgs"].append({
                            "type": "warn",
                            "text": "删除失败<br>[ %s=%s ] 人员信息不存在" % (
                                user_info_field_headers["campus_idno"], campus_idno
                            )
                        })

                else:

                    # 更新
                    if user_info:

                        # 校验字段 特定字段 锁定字段 (自定义)
                        if modifykeys is not None:
                            update_info["modifykeys"] = modifykeys
                        # 校验字段 特定字段 锁定字段 (出现就锁定)
                        else:
                            update_info["modifykeys"] = sorted(list(set(user_info["modifykeys"]+list(update_info.keys()))))

                        # 确认更新
                        update_info = {field: value for field, value in update_info.items() if value != user_info[field]}

                        # 无需更新
                        if not update_info:
                            result["counts"]["skip"] += 1
                            continue

                        # 更新信息
                        update_info["modifytime"] = datetime.datetime.now()
                        update_info["modifyuser"] = current_user.idno
                        update_result = mongo.coll_user_info.update_one(
                            {"campus_idno": campus_idno},
                            {"$set": update_info}
                        )
                        if update_result.modified_count == 1:
                            result["counts"]["update"] += 1
                        else:
                            result["counts"]["error"] += 1
                            result["msgs"].append({
                                "type": "error",
                                "text": "更新失败<br>[ %s=%s ]" % (
                                    user_info_field_headers["campus_idno"], campus_idno
                                )
                            })
                            print("[ERROR] update_user_info_by_file [ campus_idno=%s ] modified_count: %s ]\n%s" % (
                                campus_idno,
                                update_result.modified_count,
                                update_info
                            ))

                    # 新增
                    else:

                        # 获取 新增信息
                        insert_info = self.new_doc
                        insert_info["campus_idno"] = campus_idno

                        # 校验字段 特定字段 锁定字段 (自定义)
                        if modifykeys is not None:
                            update_info["modifykeys"] = modifykeys
                        # 校验字段 特定字段 锁定字段 (出现就锁定)
                        else:
                            update_info["modifykeys"] = sorted(list(set(update_info.keys())))

                        # 新增信息
                        insert_info.update(update_info)
                        insert_info["modifytime"] = datetime.datetime.now()
                        insert_info["modifyuser"] = current_user.idno
                        mongo.coll_user_info.insert_one(insert_info)
                        result["counts"]["insert"] += 1

        result["msgs"] = [{
            "type": "success",
            "text": "更新成功<br>人员信息 [ 总数=%s, 更新=%s, 新增=%s, 跳过=%s, 删除=%s, 错误=%s]" % (
                result["counts"]["total"],
                result["counts"]["update"],
                result["counts"]["insert"],
                result["counts"]["skip"],
                result["counts"]["delete"],
                result["counts"]["error"],
            )
        }] + result["msgs"]

        print("[INFO] total: %s, update: %s, insert: %s, skip: %s, delete: %s, error: %s" % (
            result["counts"]["total"],
            result["counts"]["update"],
            result["counts"]["insert"],
            result["counts"]["skip"],
            result["counts"]["delete"],
            result["counts"]["error"],
        ))

        return result

    # ---
    # function.save
    # ---

    def save_infos(self, user_infos):
        result = {
            "msgs": [],
            "file_info": None,
        }

        file_dir = "temp_dir"
        date_info = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = "user_utils-save_infos-%s.xlsx" % date_info
        save_name = "人员信息-%s.xlsx" % date_info
        file_path = file_utils.get_path(file_dir=file_dir, file_name=file_name)

        try:
            wb = Workbook()
            ws = wb.worksheets[0]

            field_headers = self.field_headers
            header_fields = {header: field for field, header in field_headers.items()}

            alignment = Alignment(
                vertical="center",
                horizontal="center"
            )

            aligntments = {header: alignment for header in header_fields}

            fonts = {
                "header": Font(
                    bold=True
                )
            }

            for col, _ in enumerate(header_fields):
                ws.column_dimensions[get_column_letter(col+1)].width = 20

            for col, header in enumerate(header_fields, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.alignment = aligntments[header]
                cell.font = fonts["header"]
                cell.number_format = numbers.FORMAT_TEXT

            for row, user_info in enumerate(user_infos, start=2):
                for col, header in enumerate(header_fields, start=1):
                    field = header_fields[header]
                    cell = ws.cell(row=row, column=col)
                    if header_fields[header] == "modifykeys":
                        cell.value = ",".join([field_headers[key] for key in user_info["modifykeys"]])
                    else:
                        cell.value = str(user_info[field]).strip()
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.alignment = aligntments[header]

            wb.save(file_path)

        except Exception as e:
            print(traceback.format_exc())
            result["msgs"].append({
                "type": "error",
                "text": "保存文件失败<br>[ file_dir=%s, file_name=%s ]" % (file_dir, file_name)
            })
            return result

        result["file_info"] = {
            "file_dir": file_dir,
            "file_name": file_name,
            "save_name": save_name
        }
        return result


user_utils = UserUtils()
