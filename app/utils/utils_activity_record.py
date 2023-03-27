"""
Utils Activity Record

# MongoDB 索引
mongo.coll_activity_record.create_index(
    name="index", unique=True, keys=[("activity_code", 1), ("campus_idno", 1)]
)
mongo.coll_activity_record.create_index(
    name="index1", keys=[("activity_code", 1)]
)
mongo.coll_activity_record.create_index(
    name="index2", keys=[("campus_idno", 1)]
)

"""

import datetime
import traceback

from flask_login import current_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from .utils_file import file_utils
from .utils_mongo import mongo


class ActivityRecordUtils(object):

    # ---
    # property
    # ---

    @property
    def new_doc(self):
        __new_doc = {
            "activity_code": "/",
            "campus_idno":   "/",
            "count":         0,

            "signup":        "/",
            "signin_record": "/",
            "takeoff":       "/",
            "score":         "/",
            "grade":         "/",
            "note":          "/",

            "createtime":    datetime.datetime.now(),
            "modifytime":    "/",
            "modifyuser":    "/",
        }
        return __new_doc

    @property
    def field_headers(self):
        __field_headers = {
            "activity_code": "活动代码", # unique with "campus_idno"   & link to "activity"
            "campus_idno":   "校园卡号", # unique with "activity_code" & link to "user"
            "count":         "数量",     # important with "course_record.activity_done"

            "signup":        "报名",
            "signin_record": "签到记录",
            "takeoff":       "请假",
            "score":         "分数",
            "grade":         "成绩",
            "note":          "备注",

            "createtime":    "创建时间",
            "modifytime":    "修改时间",
            "modifyuser":    "修改用户",
        }
        return __field_headers

    @property
    def field_options(self):
        __field_options = {
            "activity_code": sorted(mongo.coll_activity_record.distinct("activity_code")),
        }
        return __field_options

    @property
    def field_limits(self):
        __field_limits = {
        }
        return __field_limits

    @property
    def field_fixeds(self):
        __field_fixeds = [
            "createtime",
            "modifytime",
            "modifyuser",
        ]
        return __field_fixeds

    # ---
    # function.get
    # ---

    def get_record(self, activity_code, campus_idno):
        result = {
            "msgs": [],
            "activity_record": None,
        }
        # 查询活动记录
        activity_records = list(mongo.coll_activity_record.aggregate([
            {"$match": {"activity_code": activity_code, "campus_idno": campus_idno}},
            # 查询人员信息
            {"$lookup": {
                "from": "coll_user_info", "as": "user_info",
                "localField": "campus_idno", "foreignField": "campus_idno"
            }},
            {"$unwind": "$user_info"},
            # 查询活动信息
            {"$lookup": {
                "from": "coll_activity_info", "as": "activity_info",
                "localField": "activity_code", "foreignField": "activity_code"
            }},
            {"$unwind": "$activity_info"},
            # 查询课程信息
            {"$lookup": {
                "from": "coll_course_info", "as": "course_info",
                "localField": "course_code", "foreignField": "activity_info.course_code"
            }},
            {"$unwind": "$course_info"},
            # 限制数量
            {"$limit": 1}
        ]))
        if not activity_records:
            return result
        for activity_record in activity_records:
            if activity_record["count"] != 0:
                activity_record["status"] = "已完成"
            else:
                activity_record["status"] = "未开始"
        result["activity_record"] = activity_records[0]
        return result

    # ---
    # function.update
    # ---

    def update_record_by_file(self, update_file):
        from .utils_course_record import course_record_utils

        result = {
            "msgs": [],
            "counts": {
                "total":    0,
                "update":   0,
                "insert":   0,
                "skip":     0,
                "delete":   0,
                "error":    0
            }
        }

        # 读取原始数据
        wb = load_workbook(update_file, data_only=True)
        ws = wb.worksheets[0]
        rows = []
        for row in ws.iter_rows():
            rows.append([cell.value for cell in row])

        row_header = rows[0]
        row_values = rows[1:]

        activity_record_field_headers = self.field_headers
        activity_record_field_limits  = self.field_limits
        activity_record_field_fixeds  = self.field_fixeds

        activity_record_header_fields = {header: field for field, header in activity_record_field_headers.items()}

        # 校验字段 定位字段 活动代码
        if activity_record_field_headers["activity_code"] not in row_header:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少定位字段 [ %s ]" % activity_record_field_headers["activity_code"]
            })
            return result
        # 校验字段 定位字段 校园卡号
        if activity_record_field_headers["campus_idno"] not in row_header:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少定位字段 [ %s ]" % activity_record_field_headers["campus_idno"]
            })
            return result

        # 校验字段 无效字段
        invalid_headers = []
        for __header in row_header:
            if __header \
            and activity_record_header_fields[__header] in activity_record_field_headers \
            and (
                activity_record_header_fields[__header] == "activity_code" or \
                activity_record_header_fields[__header] not in activity_record_field_fixeds
            ):
                continue
            else:
                invalid_headers.append(__header)
                result["msgs"].append({
                    "type": "warn",
                    "text": "校验失败<br>存在无效字段 [ %s ]" % __header
                })

        # 校验字段 更新字段
        if len(row_header) - len(invalid_headers) <= 1:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少更新字段"
            })
            return result

        with tqdm(desc="[INFO] update_activity_record_by_file", total=len(row_values)) as pbar:

            for row_value in row_values:
                result["counts"]["total"] += 1
                pbar.update(1)

                # 获取更新信息
                delete        = False # 删除
                activity_code = None  # 活动代码
                campus_idno   = None  # 校园卡号

                update_info = {}
                default_doc = self.new_doc
                for idx, value in enumerate(row_value):
                    header = row_header[idx]
                    # 转换格式
                    if isinstance(value, (int, float)):
                        value = str(value)
                    # 提取字段 删除
                    if header == "删除":
                        if str(value).upper() in ["TRUE", "T"]:
                            delete = True
                            continue
                    # 校验字段 未知字段 跳过
                    if header not in activity_record_header_fields:
                        continue
                    # 转义字段
                    field = activity_record_header_fields[header]
                    # 提取字段 活动代码
                    if field == "activity_code":
                        activity_code = value
                        continue
                    # 提取字段 活动代码
                    if field == "campus_idno":
                        campus_idno = value
                        continue
                    # 校验字段 固定字段 跳过
                    if field in activity_record_field_fixeds:
                        continue
                    # 校验字段 所有字段 默认值
                    if not value:
                        value = default_doc[field]
                    # 校验字段 特定字段 数量
                    if field == "count":
                        value = int(value)
                    # 校验字段 限制字段
                    if field in activity_record_field_limits and value not in activity_record_field_limits[field]:
                        result["msgs"].append({
                            "type": "error",
                            "text": "校验失败<br>值不可选 [ %s=%s ]" % (activity_record_field_headers[field], value)
                        })
                        continue
                    # 记录字段
                    update_info[field] = value

                # 校验字段 定位字段 活动代码
                if mongo.coll_activity_info.count_documents({"activity_code": activity_code}) == 0:
                    result["counts"]["error"] += 1
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值不存在 [ %s=%s ]" % (
                            activity_record_field_headers["activity_code"], activity_code
                        )
                    })
                    continue
                # 校验字段 定位字段 校园卡号
                if mongo.coll_user_info.count_documents({"campus_idno": campus_idno}) == 0:
                    result["counts"]["error"] += 1
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值不存在 [ %s=%s ]" % (
                            activity_record_field_headers["campus_idno"], campus_idno
                        )
                    })
                    continue

                activity_record = mongo.coll_activity_record.find_one({"activity_code": activity_code, "campus_idno": campus_idno})

                # 删除
                if delete:

                    if activity_record:
                        delete_result = mongo.coll_activity_record.delete_one({"activity_code": activity_code, "campus_idno": campus_idno})
                        result["counts"]["delete"] += 1
                        result["msgs"].append({
                            "type": "warn",
                            "text": "删除成功<br>[ %s=%s, %s=%s ] 活动记录: %s" % (
                                activity_record_field_headers["activity_code"], activity_code,
                                activity_record_field_headers["campus_idno"],   campus_idno,
                                delete_result.deleted_count,
                            )
                        })

                    else:
                        result["counts"]["skip"] += 1
                        result["msgs"].append({
                            "type": "warn",
                            "text": "删除失败<br>[ %s=%s, %s=%s ] 活动记录不存在" % (
                                activity_record_field_headers["activity_code"], activity_code,
                                activity_record_field_headers["campus_idno"],   campus_idno,
                            )
                        })

                else:

                    # 更新
                    if activity_record:

                        # 确认更新
                        update_info = {field: value for field, value in update_info.items() if value != activity_record[field]}

                        # 无需更新
                        if not update_info:
                            result["counts"]["skip"] += 1
                            continue

                        # 更新信息
                        update_info["modifytime"] = datetime.datetime.now()
                        update_info["modifyuser"] = current_user.idno
                        update_result = mongo.coll_activity_record.update_one(
                            {"activity_code": activity_code, "campus_idno": campus_idno},
                            {"$set": update_info}
                        )
                        if update_result.modified_count == 1:
                            result["counts"]["update"] += 1

                            # 同步更新
                            if "count" in update_info:
                                # 同步更新 数量 => 课程记录
                                __activity_info = mongo.coll_activity_info.find_one(
                                    {"activity_code": activity_code}
                                )
                                __result = course_record_utils.update_record(
                                    course_code=__activity_info["course_code"],
                                    campus_idno=campus_idno
                                )
                                if __result["msgs"][-1]["type"] == "error":
                                    print("[ERROR] %s" % __result["msgs"][-1]["text"])

                        else:
                            result["counts"]["error"] += 1
                            result["msgs"].append({
                                "type": "error",
                                "text": "更新失败<br>[ %s=%s, %s=%s ]" % (
                                    activity_record_field_headers["activity_code"], activity_code,
                                    activity_record_field_headers["campus_idno"],   campus_idno,
                                )
                            })
                            print("[ERROR] update_activity_record_by_file [ activity_code=%s, modified_count: %s ]\n%s" % (
                                activity_code,
                                update_result.modified_count,
                                update_info
                            ))

                    # 新增
                    else:

                        # 获取 新增信息
                        insert_info = self.new_doc
                        insert_info["activity_code"] = activity_code
                        insert_info["campus_idno"] = campus_idno

                        # 新增信息
                        insert_info.update(update_info)
                        insert_info["modifytime"] = datetime.datetime.now()
                        insert_info["modifyuser"] = current_user.idno
                        mongo.coll_activity_record.insert_one(insert_info)
                        result["counts"]["insert"] += 1

                        # 同步更新
                        if insert_info["count"]:
                            # 同步更新 数量 => 课程记录
                            __activity_info = mongo.coll_activity_info.find_one(
                                {"activity_code": activity_code}
                            )
                            __result = course_record_utils.update_record(
                                course_code=__activity_info["course_code"],
                                campus_idno=campus_idno
                            )
                            if __result["msgs"][-1]["type"] == "error":
                                print("[ERROR] %s" % __result["msgs"][-1]["text"])

        result["msgs"] = [{
            "type": "success",
            "text": "更新成功<br>活动记录 [ 总数=%s, 更新=%s, 新增=%s, 跳过=%s, 删除=%s, 错误=%s]" % (
                result["counts"]["total"],
                result["counts"]["update"],
                result["counts"]["insert"],
                result["counts"]["skip"],
                result["counts"]["delete"],
                result["counts"]["error"],
            )
        }] + result["msgs"]

        print("[INFO] total: %s, update: %s, insert: %s, skip: %s, delete: %s, error: %s" % (
            result["counts"]["total"],
            result["counts"]["update"],
            result["counts"]["insert"],
            result["counts"]["skip"],
            result["counts"]["delete"],
            result["counts"]["error"],
        ))

        return result

    # ---
    # function.save
    # ---

    def save_records(self, activity_records):
        result = {
            "msgs": [],
            "file_info": None,
        }

        file_dir = "temp_dir"
        date_info = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = "activity_record_utils-save_records-%s.xlsx" % date_info
        save_name = "活动记录-%s.xlsx" % date_info
        file_path = file_utils.get_path(file_dir=file_dir, file_name=file_name)

        try:
            wb = Workbook()
            ws = wb.worksheets[0]

            header_fields = {header: field for field, header in self.field_headers.items()}

            alignment = Alignment(
                vertical="center",
                horizontal="center"
            )

            aligntments = {header: alignment for header in header_fields}

            fonts = {
                "header": Font(
                    bold=True
                )
            }

            for col, _ in enumerate(header_fields):
                ws.column_dimensions[get_column_letter(col+1)].width = 20

            for col, header in enumerate(header_fields, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.alignment = aligntments[header]
                cell.font = fonts["header"]
                cell.number_format = numbers.FORMAT_TEXT

            for row, activity_record in enumerate(activity_records, start=2):
                for col, header in enumerate(header_fields, start=1):
                    field = header_fields[header]
                    cell = ws.cell(row=row, column=col)
                    cell.value = str(activity_record[field]).strip().replace("'", '"')
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.alignment = aligntments[header]

            wb.save(file_path)

        except Exception as e:
            print(traceback.format_exc())
            result["msgs"].append({
                "type": "error",
                "text": "保存文件失败<br>[ file_dir=%s, file_name=%s ]" % (file_dir, file_name)
            })
            return result

        result["file_info"] = {
            "file_dir": file_dir,
            "file_name": file_name,
            "save_name": save_name
        }
        return result

    def save_records_with_detail(self, activity_records):
        from .utils_user import user_utils

        result = {
            "msgs": [],
            "file_info": None,
        }

        file_dir = "temp_dir"
        date_info = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = "activity_record_utils-save_records_with_detail-%s.xlsx" % date_info
        save_name = "活动记录[详细]-%s.xlsx" % date_info
        file_path = file_utils.get_path(file_dir=file_dir, file_name=file_name)

        try:
            wb = Workbook()
            ws = wb.worksheets[0]

            user_info_field_headers = user_utils.field_headers
            for __field in [
                "idno",
                "sex",
                "bankacc",
                "phoneno",

                "createtime",
                "modifytime",
                "modifyuser",

                "modifykeys",
            ]:
                user_info_field_headers.pop(__field)
            user_header_fields = {header: field for field, header in user_info_field_headers.items()}

            activity_record_field_headers = self.field_headers
            for __field in [
                "campus_idno",

                "createtime",
                "modifytime",
                "modifyuser",
            ]:
                activity_record_field_headers.pop(__field)
            activity_record_header_fields = {header: field for field, header in activity_record_field_headers.items()}

            alignment = Alignment(
                vertical="center",
                horizontal="center"
            )

            aligntments = {}
            aligntments.update({header: alignment for header in user_header_fields})
            aligntments.update({header: alignment for header in activity_record_header_fields})

            fonts = {
                "header": Font(
                    bold=True
                )
            }

            # 列宽
            for col, _ in enumerate(user_header_fields):
                ws.column_dimensions[get_column_letter(col+1)].width = 20
            for col, _ in enumerate(activity_record_header_fields, start=len(user_header_fields)):
                ws.column_dimensions[get_column_letter(col+1)].width = 20

            # 标题
            for col, header in enumerate(user_header_fields, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.alignment = aligntments[header]
                cell.font = fonts["header"]
                cell.number_format = numbers.FORMAT_TEXT
            for col, header in enumerate(activity_record_header_fields, start=len(user_header_fields)+1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.alignment = aligntments[header]
                cell.font = fonts["header"]
                cell.number_format = numbers.FORMAT_TEXT

            # 内容
            for row, activity_record in enumerate(activity_records, start=2):
                for col, header in enumerate(user_header_fields, start=1):
                    field = user_header_fields[header]
                    cell = ws.cell(row=row, column=col)
                    cell.value = str(activity_record["user_info"][field]).strip().replace("'", '"')
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.alignment = aligntments[header]
                for col, header in enumerate(activity_record_header_fields, start=len(user_header_fields)+1):
                    field = activity_record_header_fields[header]
                    cell = ws.cell(row=row, column=col)
                    cell.value = str(activity_record[field]).strip().replace("'", '"')
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.alignment = aligntments[header]

            wb.save(file_path)

        except Exception as e:
            print(traceback.format_exc())
            result["msgs"].append({
                "type": "error",
                "text": "保存文件失败<br>[ file_dir=%s, file_name=%s ]" % (file_dir, file_name)
            })
            return result

        result["file_info"] = {
            "file_dir": file_dir,
            "file_name": file_name,
            "save_name": save_name
        }
        return result


activity_record_utils = ActivityRecordUtils()
