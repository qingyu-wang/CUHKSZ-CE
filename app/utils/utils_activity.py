"""
Utils Activity

# MongoDB 索引
mongo.coll_activity_info.create_index(
    name="index", unique=True, keys=[("activity_code", 1)]
)

"""

import datetime
import traceback

from flask_login import current_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from .utils_file import file_utils
from .utils_mongo import mongo


class ActivityUtils(object):

    # ---
    # property
    # ---

    @property
    def new_doc(self):
        __new_doc = {
            "course_code":    "/",
            "activity_code":  "/",
            "activity_type":  "/",

            "activity_name":  "/",
            "activity_note":  "/",
            "activity_quota": "/",

            "activity_year":  "/",
            "activity_term":  "/",
            "activity_date":  "/",

            "createtime":     datetime.datetime.now(),
            "modifytime":     "/",
            "modifyuser":     "/",
        }
        return __new_doc

    @property
    def field_headers(self):
        __field_headers = {
            "course_code":    "课程代码", # link to "course_code"
            "activity_code":  "活动代码", # unique & link to "activity_record"
            "activity_type":  "活动类型", # important with "course.activity_rules" and "course_record.activity_done"

            "activity_name":  "活动名称",
            "activity_note":  "活动详情",
            "activity_quota": "活动名额",

            "activity_year":  "活动学年",
            "activity_term":  "活动学期",
            "activity_date":  "活动日期",

            "createtime":     "创建时间",
            "modifytime":     "修改时间",
            "modifyuser":     "修改用户",
        }
        return __field_headers

    @property
    def field_options(self):
        __field_options = {
            "course_code":   sorted(mongo.coll_activity_info.distinct("course_code")),
            "activity_code": sorted(mongo.coll_activity_info.distinct("activity_code")),
            "activity_type": sorted(mongo.coll_activity_info.distinct("activity_type")),

            "activity_year": sorted(mongo.coll_activity_info.distinct("activity_year")),
            "activity_term": sorted(mongo.coll_activity_info.distinct("activity_term")),
            "activity_date": sorted(mongo.coll_activity_info.distinct("activity_date")),
        }
        return __field_options

    @property
    def field_limits(self):
        __field_limits = {
            "course_code": sorted(mongo.coll_course_info.distinct("course_code")),
        }
        return __field_limits

    @property
    def field_fixeds(self):
        __field_fixeds = [
            "createtime",
            "modifytime",
            "modifyuser",
        ]
        return __field_fixeds

    # ---
    # function.get
    # ---

    def get_activity_code_options(self, field_filters):
        activity_code_options = sorted(mongo.coll_activity_info.distinct("activity_code", field_filters))
        return activity_code_options

    def get_info(self, activity_code):
        result = {
            "msgs": [],
            "activity_info": None,
        }
        activity_infos = list(mongo.coll_activity_info.aggregate([
            # 查询活动信息
            {"$match": {"activity_code": activity_code}},
            # 查询课程信息
            {"$lookup": {
                "from": "coll_course_info", "as": "course_info",
                "localField": "course_code", "foreignField": "course_code"
            }},
            {"$unwind": "$course_info"},
            # 限制数量
            {"$limit": 1}
        ]))
        if not activity_infos:
            return result
        result["activity_info"] = activity_infos[0]
        return result

    def get_info_with_detail(self, activity_code):
        return self.get_info_with_detail_2(activity_code)

    def get_info_with_detail_1(self, activity_code):
        result = {
            "msgs": [],
            "activity_info": None
        }
        activity_infos = list(mongo.coll_activity_info.aggregate([
            # 查询活动信息
            {"$match": {"activity_code": activity_code}},
            # 查询课程信息
            {"$lookup": {
                "from": "coll_course_info", "as": "course_info",
                "localField": "course_code", "foreignField": "course_code"
            }},
            {"$unwind": "$course_info"},
            {"$limit": 1}
        ]))
        if not activity_infos:
            return result
        result["activity_info"] = activity_infos[0]

        activity_records = list(mongo.coll_activity_record.aggregate([
            # 查询活动记录
            {"$match": {"activity_code": activity_code}},
            # 查询人员信息
            {"$lookup": {
                "from": "coll_user_info", "as": "user_info",
                "localField": "campus_idno", "foreignField": "campus_idno"
            }},
            {"$unwind": "$user_info"},
            {"$sort":{"campus_idno": 1}},
        ]))
        result["activity_info"]["activity_records"] = activity_records

        # 查询活动概况
        activity_overview = {
            "num_total":   0,
            "num_done":    0,
            "num_signup":  0,
            "num_takeoff": 0,
        }
        for activity_record in result["activity_info"]["activity_records"]:
            activity_overview["num_total"] += 1
            if activity_record["count"] != 0:
                activity_record["status"] = "已完成"
                activity_overview["num_done"] += 1
            else:
                activity_record["status"] = "未开始"
            if str(activity_record["signup"]).upper() in ["TRUE", "T"]:
                activity_overview["num_signup"] += 1
            if str(activity_record["takeoff"]).upper() in ["TRUE", "T"]:
                activity_overview["num_takeoff"] += 1
        result["activity_info"]["activity_overview"] = activity_overview
        return result

    def get_info_with_detail_2(self, activity_code):
        result = {
            "msgs": [],
            "activity_info": None,
        }
        activity_infos = list(mongo.coll_activity_info.aggregate([
            # 查询 活动信息
            {"$match": {"activity_code": activity_code}},
            # 查询 课程信息
            {"$lookup": {
                "from": "coll_course_info", "as": "course_info",
                "localField": "course_code", "foreignField": "course_code"
            }},
            {"$unwind": "$course_info"},
            {"$limit": 1},
            # 查询 活动记录
            {"$lookup": {
                "from": "coll_activity_record", "as": "activity_records",
                "let": {
                    "activity_code": "$activity_code"
                },
                "pipeline": [
                    {"$match": {"$expr": {"$and": [
                        {"$eq": ["$activity_code", "$$activity_code"]}
                    ]}}},
                    # 查询 人员信息
                    {"$lookup": {
                        "from": "coll_user_info", "as": "user_info",
                        "localField": "campus_idno", "foreignField": "campus_idno"
                    }},
                    {"$unwind": "$user_info"},
                    {"$sort":{"campus_idno": 1}},
                ]
            }}
        ]))
        if not activity_infos:
            return None
        result["activity_info"] = activity_infos[0]

        # 后处理 活动概况
        activity_overview = {
            "num_total":   0,
            "num_done":    0,
            "num_signup":  0,
            "num_takeoff": 0,
        }
        for activity_record in result["activity_info"]["activity_records"]:
            activity_overview["num_total"] += 1
            if activity_record["count"] != 0:
                activity_overview["num_done"] += 1
            if str(activity_record["signup"]).upper() in ["TRUE", "T"]:
                activity_overview["num_signup"] += 1
            if str(activity_record["takeoff"]).upper() in ["TRUE", "T"]:
                activity_overview["num_takeoff"] += 1
        result["activity_info"]["activity_overview"] = activity_overview
        return result

    # ---
    # function.update
    # ---

    def update_info(self, new_activity_info, old_activity_code):
        result = {
            "msgs": [],
        }
        old_activity_info = mongo.coll_activity_info.find_one({"activity_code": old_activity_code})

        # 获取更新信息
        update_info = {}

        activity_info_field_headers = self.field_headers
        activity_info_field_limits  = self.field_limits
        activity_info_field_fixeds  = self.field_fixeds
        default_doc = self.new_doc

        for field in new_activity_info:
            new_value = new_activity_info[field]
            old_value = old_activity_info[field]
            # 校验字段 固定字段 跳过
            if field in activity_info_field_fixeds:
                continue
            # 校验字段 所有字段 默认值
            if not new_value:
                new_value = default_doc[field]
            # 校验字段 特定字段 活动代码
            if field == "activity_code" and new_value != old_value:
                activity_num = mongo.coll_activity_info.count_documents({"activity_code": new_value})
                if activity_num != 0:
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值已存在 [ %s=%s ]" % (activity_info_field_headers[field], new_value)
                    })
                    return result
            # 校验字段 限制字段
            if field in activity_info_field_limits and new_value not in activity_info_field_limits[field]:
                result["msgs"].append({
                    "type": "error",
                    "text": "校验失败<br>值不可选 [ %s=%s ]" % (activity_info_field_headers[field], new_value)
                })
                return result
            # 更新字段
            if new_value != old_value:
                update_info[field] = new_value

        # 无需更新
        if len(update_info) == 0:
            result["msgs"].append({
                "type": "warn",
                "text": "无需更新"
            })
            return result

        # 更新信息
        update_info["modifytime"] = datetime.datetime.now()
        update_info["modifyuser"] = current_user.idno
        result_1 = mongo.coll_activity_info.update_one(
            {"activity_code": old_activity_code},
            {"$set": update_info}
        )

        if result_1.modified_count != 0:
            result["msgs"].append({
                "type": "success",
                "text": "更新成功<br>活动信息 [ 总数=1, 更新=%s, 错误=0 ]" % result_1.modified_count
            })

            # 同步更新 活动代码
            if "activity_code" in update_info:

                # 更新 活动记录
                activity_record_num = mongo.coll_activity_record.count_documents({"activity_code": old_activity_code})
                result_2 = mongo.coll_activity_record.update_many(
                    {"activity_code": old_activity_code},
                    {"$set": {"activity_code": update_info["activity_code"]}}
                )
                if result_2.modified_count == activity_record_num:
                    result["msgs"].append({
                        "type": "success",
                        "text": "更新成功<br>活动代码 => 活动记录 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                            activity_record_num,
                            result_2.modified_count,
                            activity_record_num-result_2.modified_count,
                        )
                    })
                else:
                    result["msgs"].append({
                        "type": "error",
                        "text": "更新失败<br>活动代码 => 活动记录 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                            activity_record_num,
                            result_2.modified_count,
                            activity_record_num-result_2.modified_count,
                        )
                    })

        else:
            result["msgs"].append({
                "type": "error",
                "text": "更新失败<br>活动信息 [ 总数=1, 更新=0, 错误=1 ]"
            })
        return result

    def update_info_by_file(self, update_file):
        from .utils_course_record import course_record_utils

        result = {
            "msgs": [],
            "counts": {
                "total":    0,
                "update":   0,
                "insert":   0,
                "skip":     0,
                "delete":   0,
                "error":    0
            }
        }

        # 读取原始数据
        wb = load_workbook(update_file, data_only=True)
        ws = wb.worksheets[0]
        rows = []
        for row in ws.iter_rows():
            rows.append([cell.value for cell in row])

        row_header = rows[0]
        row_values = rows[1:]

        activity_info_field_headers = self.field_headers
        activity_info_field_limits  = self.field_limits
        activity_info_field_fixeds  = self.field_fixeds

        activity_info_header_fields = {header: field for field, header in activity_info_field_headers.items()}

        # 校验字段 定位字段 活动代码
        if activity_info_field_headers["activity_code"] not in row_header:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少定位字段 [ %s ]" % activity_info_field_headers["activity_code"]
            })
            return result

        # 校验字段 无效字段
        invalid_headers = []
        for __header in row_header:
            if __header == "删除":
                continue
            elif __header \
            and activity_info_header_fields[__header] in activity_info_field_headers \
            and (
                activity_info_header_fields[__header] == "activity_code" or \
                activity_info_header_fields[__header] == "campus_idno" or \
                activity_info_header_fields[__header] not in activity_info_field_fixeds
            ):
                continue
            else:
                invalid_headers.append(__header)
                result["msgs"].append({
                    "type": "warn",
                    "text": "校验失败<br>存在无效字段 [ %s ]" % __header
                })

        # 校验字段 更新字段
        if len(row_header) - len(invalid_headers) <= 1:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少更新字段"
            })
            return result

        with tqdm(desc="[INFO] update_activity_info_by_file", total=len(row_values)) as pbar:

            for row_value in row_values:
                result["counts"]["total"] += 1
                pbar.update(1)

                # 获取更新信息
                delete        = False # 删除
                activity_code = None  # 活动代码

                update_info = {}
                default_doc = self.new_doc
                for idx, value in enumerate(row_value):
                    header = row_header[idx]
                    # 提取字段 删除
                    if header == "删除":
                        if str(value).upper() in ["TRUE", "T"]:
                            delete = True
                            continue
                    # 校验字段 未知字段 跳过
                    if header not in activity_info_header_fields:
                        continue
                    # 转义字段
                    field = activity_info_header_fields[header]
                    # 提取字段 活动代码
                    if field == "activity_code":
                        activity_code = value
                        continue
                    # 校验字段 固定字段 跳过
                    if field in activity_info_field_fixeds:
                        continue
                    # 校验字段 所有字段 默认值
                    if not value:
                        value = default_doc[field]
                    # 校验字段 限制字段
                    if field in activity_info_field_limits and value not in activity_info_field_limits[field]:
                        result["msgs"].append({
                            "type": "error",
                            "text": "校验失败<br>值不可选 [ %s=%s ]" % (activity_info_field_headers[field], value)
                        })
                        continue
                    # 记录字段
                    update_info[field] = value

                # 校验字段 定位字段 活动代码
                if not activity_code:
                    result["counts"]["error"] += 1
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值不能为空 [ %s ]" % activity_info_field_headers["activity_code"]
                    })
                    continue

                activity_info = mongo.coll_activity_info.find_one({"activity_code": activity_code})

                # 删除
                if delete:

                    if activity_info:
                        # 有"活动记录"不能删除"活动信息"
                        if mongo.coll_activity_record.count_documents({"activity_code": activity_code}) != 0:
                            result["counts"]["error"] += 1
                            result["msgs"].append({
                                "type": "error",
                                "text": "删除失败<br>[ %s=%s ] 存在活动记录" % (
                                    activity_info_field_headers["activity_code"], activity_code
                                )
                            })

                        else:
                            delete_result = mongo.coll_activity_info.delete_one({"activity_code": activity_code})
                            result["counts"]["delete"] += 1
                            result["msgs"].append({
                                "type": "warn",
                                "text": "删除成功<br>[ %s=%s ] 活动信息: %s" % (
                                    activity_info_field_headers["activity_code"], activity_code,
                                    delete_result.deleted_count,
                                )
                            })

                    else:
                        result["counts"]["skip"] += 1
                        result["msgs"].append({
                            "type": "warn",
                            "text": "删除失败<br>[ %s=%s ] 活动信息不存在" % (
                                activity_info_field_headers["activity_code"], activity_code
                            )
                        })

                else:

                    # 更新
                    if activity_info:

                        # 确认更新
                        update_info = {field: value for field, value in update_info.items() if value != activity_info[field]}

                        # 无需更新
                        if not update_info:
                            result["counts"]["skip"] += 1
                            continue

                        # 更新信息
                        update_info["modifytime"] = datetime.datetime.now()
                        update_info["modifyuser"] = current_user.idno
                        update_result = mongo.coll_activity_info.update_one(
                            {"activity_code": activity_code},
                            {"$set": update_info}
                        )
                        if update_result.modified_count == 1:
                            result["counts"]["update"] += 1

                            # 同步更新
                            if "course_code" in update_info:
                                # 同步更新 课程代码 => 课程记录
                                course_record_num = 0
                                course_record_modified_count = 0
                                with tqdm(desc="[INFO] update_course_record") as pbar:
                                    for __activity_record in mongo.coll_activity_record.find(
                                        {"activity_code": activity_code}
                                    ):
                                        for __course_record in mongo.coll_course_record.find(
                                            {"campus_idno": __activity_record["campus_idno"]}
                                        ):
                                            __result = course_record_utils.update_record(
                                                course_code=__course_record["course_code"],
                                                campus_idno=__course_record["campus_idno"]
                                            )
                                            pbar.update(1)
                                            course_record_num += 1
                                            if __result["msgs"][-1]["type"] == "success":
                                                course_record_modified_count += 1
                                if course_record_modified_count == course_record_num:
                                    result["msgs"].append({
                                        "type": "success",
                                        "text": "更新成功<br>课程代码 => 课程记录 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                                            course_record_num,
                                            course_record_modified_count,
                                            course_record_num - course_record_modified_count
                                        )
                                    })
                                else:
                                    result["msgs"].append({
                                        "type": "error",
                                        "text": "更新失败<br>课程代码 => 课程记录 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                                            course_record_num,
                                            course_record_modified_count,
                                            course_record_num - course_record_modified_count
                                        )
                                    })

                            elif "activity_type" in update_info:
                                # 同步更新 活动类型 => 课程记录
                                course_record_num = 0
                                course_record_modified_count = 0
                                with tqdm(desc="[INFO] update_course_record") as pbar:
                                    for __activity_record in mongo.coll_activity_record.find(
                                        {"activity_code": activity_code}
                                    ):
                                        for __course_record in mongo.coll_course_record.find(
                                            {
                                                "course_code": activity_info["course_code"],
                                                "campus_idno": __activity_record["campus_idno"]
                                            }
                                        ):
                                            __result = course_record_utils.update_record(
                                                course_code=__course_record["course_code"],
                                                campus_idno=__course_record["campus_idno"]
                                            )
                                            pbar.update(1)
                                            course_record_num += 1
                                            if __result["msgs"][-1]["type"] == "success":
                                                course_record_modified_count += 1
                                if course_record_modified_count == course_record_num:
                                    result["msgs"].append({
                                        "type": "success",
                                        "text": "更新成功<br>活动规则 => 课程记录 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                                            course_record_num,
                                            course_record_modified_count,
                                            course_record_num - course_record_modified_count
                                        )
                                    })
                                else:
                                    result["msgs"].append({
                                        "type": "error",
                                        "text": "更新失败<br>活动规则 => 课程记录 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                                            course_record_num,
                                            course_record_modified_count,
                                            course_record_num - course_record_modified_count
                                        )
                                    })

                        else:
                            result["counts"]["error"] += 1
                            result["msgs"].append({
                                "type": "error",
                                "text": "更新失败<br>[ %s=%s ]" % (
                                    activity_info_field_headers["activity_code"], activity_code
                                )
                            })
                            print("[ERROR] update_activity_info_by_file [ activity_code=%s ] modified_count: %s ]\n%s" % (
                                activity_code,
                                update_result.modified_count,
                                update_info
                            ))

                    # 新增
                    else:

                        # 获取 新增信息
                        insert_info = self.new_doc
                        insert_info["activity_code"] = activity_code

                        # 新增信息
                        insert_info.update(update_info)
                        insert_info["modifytime"] = datetime.datetime.now()
                        insert_info["modifyuser"] = current_user.idno
                        mongo.coll_activity_info.insert_one(insert_info)
                        result["counts"]["insert"] += 1

        result["msgs"] = [{
            "type": "success",
            "text": "更新成功<br>活动信息 [ 总数=%s, 更新=%s, 新增=%s, 跳过=%s, 删除=%s, 错误=%s]" % (
                result["counts"]["total"],
                result["counts"]["update"],
                result["counts"]["insert"],
                result["counts"]["skip"],
                result["counts"]["delete"],
                result["counts"]["error"],
            )
        }] + result["msgs"]

        print("[INFO] total: %s, update: %s, insert: %s, skip: %s, delete: %s, error: %s" % (
            result["counts"]["total"],
            result["counts"]["update"],
            result["counts"]["insert"],
            result["counts"]["skip"],
            result["counts"]["delete"],
            result["counts"]["error"],
        ))

        return result

    # ---
    # function.create
    # ---

    def create_info(self, new_activity_info):
        result = {
            "msgs": [],
        }

        # 获取 新增信息
        insert_info = self.new_doc

        activity_info_field_headers = self.field_headers
        activity_info_field_limits  = self.field_limits
        activity_info_field_fixeds  = self.field_fixeds
        default_doc = self.new_doc

        for field in new_activity_info:
            new_value = new_activity_info[field]
            # 校验字段 固定字段 跳过
            if field in activity_info_field_fixeds:
                continue
            # 校验字段 所有字段 默认值
            if not new_value:
                new_value = default_doc[field]
            # 校验字段 特定字段 活动代码
            if field == "activity_code":
                activity_num = mongo.coll_activity_info.count_documents({"activity_code": new_value})
                if activity_num != 0:
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值已存在 [ %s=%s ]" % (activity_info_field_headers[field], new_value)
                    })
                    return result
            # 校验字段 限制字段
            if field in activity_info_field_limits and new_value not in activity_info_field_limits[field]:
                result["msgs"].append({
                    "type": "error",
                    "text": "校验失败<br>值不可选 [ %s=%s ]" % (activity_info_field_headers[field], new_value)
                })
                return result
            # 新增字段
            insert_info[field] = new_value

        # 新增信息
        insert_info["modifytime"] = datetime.datetime.now()
        insert_info["modifyuser"] = current_user.idno
        mongo.coll_activity_info.insert_one(insert_info)
        result["msgs"].append({
            "type": "success",
            "text": "新增成功<br>活动信息 [ 总数=1, 新增=1, 错误=0 ]"
        })
        return result

    # ---
    # function.save
    # ---

    def save_infos(self, activity_infos):
        result = {
            "msgs": [],
            "file_info": None,
        }

        file_dir = "temp_dir"
        date_info = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = "activity_utils-save_infos-%s.xlsx" % date_info
        save_name = "活动信息-%s.xlsx" % date_info
        file_path = file_utils.get_path(file_dir=file_dir, file_name=file_name)

        try:
            wb = Workbook()
            ws = wb.worksheets[0]

            header_fields = {header: field for field, header in self.field_headers.items()}

            alignment = Alignment(
                vertical="center",
                horizontal="center"
            )

            aligntments = {header: alignment for header in header_fields}

            fonts = {
                "header": Font(
                    bold=True
                )
            }

            for col, _ in enumerate(header_fields):
                ws.column_dimensions[get_column_letter(col+1)].width = 20

            for col, header in enumerate(header_fields, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.alignment = aligntments[header]
                cell.font = fonts["header"]
                cell.number_format = numbers.FORMAT_TEXT

            for row, activity_info in enumerate(activity_infos, start=2):
                for col, header in enumerate(header_fields, start=1):
                    field = header_fields[header]
                    cell = ws.cell(row=row, column=col)
                    cell.value = str(activity_info[field]).strip().replace("'", '"')
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.alignment = aligntments[header]

            wb.save(file_path)

        except Exception as e:
            print(traceback.format_exc())
            result["msgs"].append({
                "type": "error",
                "text": "保存文件失败<br>[ file_dir=%s, file_name=%s ]" % (file_dir, file_name)
            })
            return result

        result["file_info"] = {
            "file_dir": file_dir,
            "file_name": file_name,
            "save_name": save_name
        }
        return result


activity_utils = ActivityUtils()
