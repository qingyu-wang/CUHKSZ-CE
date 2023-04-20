"""
Utils Course

# MongoDB 索引
mongo.coll_course_info.create_index(
    name="index", unique=True, keys=[("course_code", 1)]
)

"""

import datetime
import json
import traceback

from flask_login import current_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from .utils_file import file_utils
from .utils_mongo import mongo


class CourseUtils(object):

    # ---
    # property
    # ---

    @property
    def new_doc(self):
        __new_doc = {
            "course_code":    "/",
            "activity_rules": "[]",

            "course_name":    "/",

            "createtime":     datetime.datetime.now(),
            "modifytime":     "/",
            "modifyuser":     "/",
        }
        return __new_doc

    @property
    def field_headers(self):
        __field_headers = {
            "course_code":    "课程代码", # unique & link to "activity" and "course_record"
            "activity_rules": "活动规则", # important with "course_record.acivity_done"

            "course_name":    "课程名称",

            "createtime":     "创建时间",
            "modifytime":     "修改时间",
            "modifyuser":     "修改用户",
        }
        return __field_headers

    @property
    def field_options(self):
        __field_options = {
            "course_code": sorted(mongo.coll_course_info.distinct("course_code")),
        }
        return __field_options

    @property
    def field_limits(self):
        __field_limits = {
        }
        return __field_limits

    @property
    def field_fixeds(self):
        __field_fixeds = [
            "createtime",
            "modifytime",
            "modifyuser",
        ]
        return __field_fixeds

    # ---
    # function.get
    # ---

    def get_info(self, course_code, field_filters={}):
        return self.get_info_2(course_code, field_filters)

    def get_info_1(self, course_code, field_filters={}):
        result = {
            "msgs": [],
            "course_info": None,
        }

        # 查询课程信息
        course_info = mongo.coll_course_info.find_one({"course_code": course_code})
        if course_info is None:
            return result
        # 查询活动信息
        field_filters["course_code"] = course_code
        course_info["activity_infos"] = list(mongo.coll_activity_info.find(field_filters).sort([("activity_code", -1)]))

        result["course_info"] = course_info
        return result

    def get_info_2(self, course_code, field_filters={}):
        result = {
            "msgs": [],
            "course_info": None,
        }
        course_infos = list(mongo.coll_course_info.aggregate([
            # 查询课程信息
            {"$match": {"course_code": course_code}},
            # 查询活动信息
            {"$lookup": {
                "from": "coll_activity_info", "as": "activity_infos",
                "let": {
                    "course_code": "$course_code"
                },
                "pipeline": [
                    {"$match": {"$expr": {"$and":
                        [{"$eq": ["$course_code", "$$course_code"]}] +
                        [{"$eq": ["$%s" % key, val]} for key, val in field_filters.items()]
                    }}},
                    {"$sort":{"activity_code": 1}},
                ]
            }}
        ]))
        if not course_infos:
            return result
        result["course_info"] = course_infos[0]
        return result

    def get_info_with_detail(self, course_code, field_filters):
        from .utils_course_record import course_record_utils

        result = {
            "msgs": [],
            "course_info": None,
        }

        # 查询校园卡号 (人员信息)
        campus_idnos_1 = [i["campus_idno"] for i in mongo.coll_user_info.find(
            field_filters["user_info"], {"_id": 0, "campus_idno": 1}
        )]
        # 查询校园卡号 (课程记录)
        campus_idnos_2 = [i["campus_idno"] for i in mongo.coll_course_record.find(
            {"course_code": course_code}, {"_id": 0, "campus_idno": 1}
        )]
        print("[INFO] 校园卡号数量 (人员信息): %8d" % len(campus_idnos_1))
        print("[INFO] 校园卡号数量 (课程记录): %8d" % len(campus_idnos_2))

        # 新增课程记录
        for campus_idno in campus_idnos_1:
            if campus_idno not in campus_idnos_2:
                __result = course_record_utils.create_record(
                    course_code=course_code,
                    campus_idno=campus_idno
                )
                result["msgs"].extend(__result["msgs"])
                __result = course_record_utils.update_record(
                    course_code=course_code,
                    campus_idno=campus_idno
                )
                result["msgs"].extend(__result["msgs"])

        # 查询课程信息
        course_infos = list(mongo.coll_course_info.aggregate([
            # 查询课程信息
            {"$match": {"course_code": course_code}},
            # 查询活动信息
            {"$lookup": {
                "from": "coll_course_record", "as": "course_records",
                "let": {
                    "course_code": "$course_code"
                },
                "pipeline": [
                    {"$match": {"$expr": {"$and":
                        [{"$eq": ["$course_code", "$$course_code"]}] +
                        [{"$eq": ["$%s" % key, val]} for key, val in field_filters["course_record"].items()]
                    }}},
                    {"$lookup":{
                        "from": "coll_user_info", "as": "user_info",
                        "localField": "campus_idno", "foreignField": "campus_idno"
                    }},
                    {"$unwind": "$user_info"},
                    {"$match": {"$expr": {"$and":
                        [{"$eq": ["$user_info.%s" % key, val]} for key, val in field_filters["user_info"].items()]
                    }}},
                    {"$sort":{"campus_idno": 1}},
                ]
            }}
        ]))
        if not course_infos:
            return result
        course_info = course_infos[0]

        course_overview = {
            "num_total":  0,
            "num_auth":   0,
            "num_done":   0,
            "num_doing":  0,
            "num_undone": 0,
        }
        for course_record in course_info["course_records"]:
            course_overview["num_total"]  += 1
            course_overview["num_auth"]   += 1 if course_record["status"] == "已认证" else 0
            course_overview["num_done"]   += 1 if course_record["status"] == "已完成" else 0
            course_overview["num_doing"]  += 1 if course_record["status"] == "进行中" else 0
            course_overview["num_undone"] += 1 if course_record["status"] == "未开始" else 0
        course_info["course_overview"] = course_overview

        result["course_info"] = course_info
        return result

    # ---
    # function.update
    # ---

    def update_info(self, new_course_info, old_course_code):
        from .utils_course_record import course_record_utils

        result = {
            "msgs": [],
        }
        old_course_info = mongo.coll_course_info.find_one({"course_code": old_course_code})

        # 获取 更新信息
        update_info = {}

        course_info_field_headers = self.field_headers
        course_info_field_limits  = self.field_limits
        course_info_field_fixeds  = self.field_fixeds

        default_doc = self.new_doc

        for field in new_course_info:
            new_value = new_course_info[field]
            old_value = old_course_info[field]
            # 校验字段 固定字段 跳过
            if field in course_info_field_fixeds:
                continue
            # 校验字段 所有字段 默认值
            if not new_value:
                new_value = default_doc[field]
            # 校验字段 特定字段 课程代码
            if field == "course_code" and new_value != old_value:
                course_num = mongo.coll_course_info.count_documents({"course_code": new_value})
                if course_num != 0:
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值已存在 [ %s=%s ]" % (course_info_field_headers[field], new_value)
                    })
                    return result
            # 校验字段 特定字段 活动规则
            if field == "activity_rules":
                try:
                    new_value = json.loads(new_value)
                except json.JSONDecodeError:
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值不正确 [ %s=%s ]" % (course_info_field_headers[field], new_value)
                    })
                    return result
            # 校验字段 限制字段
            if field in course_info_field_limits and new_value not in course_info_field_limits[field]:
                result["msgs"].append({
                    "type": "error",
                    "text": "校验失败<br>值不可选 [ %s=%s ]" % (course_info_field_headers[field], new_value)
                })
                return result
            # 更新字段
            if new_value != old_value:
                update_info[field] = new_value

        # 无需更新
        if len(update_info) == 0:
            result["msgs"].append({
                "type": "warn",
                "text": "无需更新"
            })
            return result

        # 更新信息
        update_info["modifytime"] = datetime.datetime.now()
        update_info["modifyuser"] = current_user.idno
        result_1 = mongo.coll_course_info.update_one(
            {"course_code": old_course_code},
            {"$set": update_info}
        )

        if result_1.modified_count != 0:
            result["msgs"].append({
                "type": "success",
                "text": "更新成功<br>课程信息 [ 总数=1, 更新=%s, 错误=0 ]" % result_1.modified_count
            })

            # 同步更新 课程代码
            if "course_code" in update_info:

                # 更新 活动信息
                activity_info_num = mongo.coll_activity_info.count_documents({"course_code": old_course_code})
                result_3 = mongo.coll_activity_info.update_many(
                    {"course_code": old_course_code},
                    {"$set": {"course_code": update_info["course_code"]}}
                )
                if result_3.modified_count == activity_info_num:
                    result["msgs"].append({
                        "type": "success",
                        "text": "更新成功<br>课程代码 => 活动信息 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                            activity_info_num,
                            result_3.modified_count,
                            activity_info_num-result_3.modified_count,
                        )
                    })
                else:
                    result["msgs"].append({
                        "type": "error",
                        "text": "更新失败<br>课程代码 => 活动信息 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                            activity_info_num,
                            result_3.modified_count,
                            activity_info_num-result_3.modified_count,
                        )
                    })

                # 更新 课程记录
                course_record_num = mongo.coll_course_record.count_documents({"course_code": old_course_code})
                result_4 = mongo.coll_course_record.update_many(
                    {"course_code": old_course_code},
                    {"$set": {"course_code": update_info["course_code"]}}
                )
                if result_4.modified_count == course_record_num:
                    result["msgs"].append({
                        "type": "success",
                        "text": "更新成功<br>课程代码 => 课程记录 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                            course_record_num,
                            result_4.modified_count,
                            course_record_num-result_4.modified_count,
                        )
                    })
                else:
                    result["msgs"].append({
                        "type": "error",
                        "text": "更新失败<br>课程代码 => 课程记录 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                            course_record_num,
                            result_4.modified_count,
                            course_record_num-result_4.modified_count,
                        )
                    })

        else:
            result["msgs"].append({
                "type": "error",
                "text": "更新失败<br>课程信息 [ 总数=1, 更新=0, 错误=1 ]"
            })
        return result

    def update_info_by_file(self, update_file):
        from .utils_course_record import course_record_utils

        result = {
            "msgs": [],
            "counts": {
                "total":    0,
                "update":   0,
                "insert":   0,
                "skip":     0,
                "delete":   0,
                "error":    0
            }
        }

        # 读取原始数据
        wb = load_workbook(update_file, data_only=True)
        ws = wb.worksheets[0]
        rows = []
        for row in ws.iter_rows():
            rows.append([cell.value for cell in row])

        row_header = rows[0]
        row_values = rows[1:]

        course_info_field_headers = self.field_headers
        course_info_field_limits  = self.field_limits
        course_info_field_fixeds  = self.field_fixeds

        course_info_header_fields = {header: field for field, header in course_info_field_headers.items()}

        # 校验字段 定位字段 课程代码
        if course_info_field_headers["course_code"] not in row_header:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少定位字段 [ %s ]" % course_info_field_headers["course_code"]
            })
            return result

        # 校验字段 无效字段
        invalid_headers = []
        for __header in row_header:
            if __header == "删除":
                continue
            elif __header \
            and course_info_header_fields[__header] in course_info_field_headers \
            and (
                course_info_header_fields[__header] == "course_code" or \
                course_info_header_fields[__header] not in course_info_field_fixeds
            ):
                continue
            else:
                invalid_headers.append(__header)
                result["msgs"].append({
                    "type": "warn",
                    "text": "校验失败<br>存在无效字段 [ %s ]" % __header
                })

        # 校验字段 更新字段
        if len(row_header) - len(invalid_headers) <= 1:
            result["msgs"].append({
                "type": "error",
                "text": "校验失败<br>缺少更新字段"
            })
            return result

        with tqdm(desc="[INFO] update_course_info_by_file", total=len(row_values)) as pbar:

            for row_value in row_values:
                result["counts"]["total"] += 1
                pbar.update(1)

                # 获取更新信息
                delete        = False # 删除
                activity_code = None  # 活动代码

                update_info = {}
                default_doc = self.new_doc
                for idx, value in enumerate(row_value):
                    header = row_header[idx]
                    # 提取字段 删除
                    if header == "删除":
                        if str(value).upper() in ["TRUE", "T"]:
                            delete = True
                            continue
                    # 校验字段 未知字段 跳过
                    if header not in course_info_header_fields:
                        continue
                    # 转义字段
                    field = course_info_header_fields[header]
                    # 提取字段 课程代码
                    if field == "course_code":
                        course_code = value
                        continue
                    # 校验字段 固定字段 跳过
                    if field in course_info_field_fixeds:
                        continue
                    # 校验字段 所有字段 默认值
                    if not value:
                        value = default_doc[field]
                    # 校验字段 特定字段 活动规则
                    if field == "activity_rules":
                        try:
                            value = json.loads(value)
                        except json.JSONDecodeError:
                            result["msgs"].append({
                                "type": "error",
                                "text": "校验失败<br>值不正确 [ %s=%s ]" % (course_info_field_headers[field], value)
                            })
                            return result
                    # 校验字段 限制字段
                    if field in course_info_field_limits and value not in course_info_field_limits[field]:
                        result["msgs"].append({
                            "type": "error",
                            "text": "校验失败<br>值不可选 [ %s=%s ]" % (course_info_field_headers[field], value)
                        })
                        continue
                    # 记录字段
                    update_info[field] = value

                # 校验字段 定位字段 课程代码
                if not course_code:
                    result["counts"]["error"] += 1
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值不能为空 [ %s ]" % course_info_field_headers["course"]
                    })
                    continue

                course_info = mongo.coll_course_info.find_one({"course_code": course_code})

                # 删除
                if delete:

                    if course_info:
                        # 有"活动信息"不能删除"课程信息"
                        if mongo.coll_activity_info.count_documents({"activity_code": activity_code}) != 0:
                            result["counts"]["error"] += 1
                            result["msgs"].append({
                                "type": "error",
                                "text": "删除失败<br>[ %s=%s ] 存在活动信息" % (
                                    course_info_field_headers["activity_code"], activity_code
                                )
                            })

                        else:
                            delete_result_1 = mongo.coll_course_info.delete_one({"course_code": course_code})
                            delete_result_2 = mongo.coll_course_record.delete_many({"course_code": course_code})
                            result["counts"]["delete"] += 1
                            result["msgs"].append({
                                "type": "success",
                                "text": "删除成功<br>[ %s=%s ] 课程信息: %s, 课程记录: %s" % (
                                    course_info_field_headers["course_code"], course_code,
                                    delete_result_1.deleted_count,
                                    delete_result_2.deleted_count,
                                )
                            })

                    else:
                        result["counts"]["skip"] += 1
                        result["msgs"].append({
                            "type": "warn",
                            "text": "删除失败<br>[ %s=%s ] 课程信息不存在" % (
                                course_info_field_headers["course_code"], course_code
                            )
                        })

                else:

                    # 更新
                    if course_info:

                        # 确认更新
                        update_info = {field: value for field, value in update_info.items() if value != course_info[field]}

                        # 无需更新
                        if not update_info:
                            result["counts"]["skip"] += 1
                            continue

                        # 更新信息
                        update_info["modifytime"] = datetime.datetime.now()
                        update_info["modifyuser"] = current_user.idno
                        update_result = mongo.coll_course_info.update_one(
                            {"course_code": course_code},
                            {"$set": update_info}
                        )
                        if update_result.modified_count == 1:
                            result["counts"]["update"] += 1

                            # 同步更新
                            if "activity_rules" in update_info:
                                # 同步更新 活动规则 => 课程记录
                                course_record_num = mongo.coll_course_record.count_documents({"course_code": course_code})
                                course_record_modified_count = 0
                                with tqdm(desc="[INFO] update_course_record", total=course_record_num) as pbar:
                                    for course_record in mongo.coll_course_record.find({"course_code": course_code}).sort([("campus_idno", 1)]):
                                        __result = course_record_utils.update_record(
                                            course_code=course_record["course_code"],
                                            campus_idno=course_record["campus_idno"]
                                        )
                                        if __result["msgs"][-1]["type"] == "success":
                                            course_record_modified_count += 1
                                        pbar.update(1)
                                if course_record_modified_count == course_record_num:
                                    result["msgs"].append({
                                        "type": "success",
                                        "text": "更新成功<br>活动规则 => 课程记录 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                                            course_record_num,
                                            course_record_modified_count,
                                            course_record_num - course_record_modified_count
                                        )
                                    })
                                else:
                                    result["msgs"].append({
                                        "type": "error",
                                        "text": "更新失败<br>活动规则 => 课程记录 [ 总数=%s, 更新=%s, 错误=%s ]" % (
                                            course_record_num,
                                            course_record_modified_count,
                                            course_record_num - course_record_modified_count
                                        )
                                    })

                        else:
                            result["counts"]["error"] += 1
                            result["msgs"].append({
                                "type": "error",
                                "text": "更新失败<br>[ %s=%s ]" % (
                                    course_info_field_headers["course_code"], course_code
                                )
                            })
                            print("[ERROR] update_course_info_by_file [ course_code=%s ] modified_count: %s ]\n%s" % (
                                course_code,
                                update_result.modified_count,
                                update_info
                            ))

                    # 新增
                    else:

                        # 获取 新增信息
                        insert_info = self.new_doc
                        insert_info["course_code"] = course_code

                        # 新增信息
                        insert_info.update(update_info)
                        insert_info["modifytime"] = datetime.datetime.now()
                        insert_info["modifyuser"] = current_user.idno
                        mongo.coll_course_info.insert_one(insert_info)
                        result["counts"]["insert"] += 1

        result["msgs"] = [{
            "type": "success",
            "text": "更新成功<br>课程信息 [ 总数=%s, 更新=%s, 新增=%s, 跳过=%s, 删除=%s, 错误=%s]" % (
                result["counts"]["total"],
                result["counts"]["update"],
                result["counts"]["insert"],
                result["counts"]["skip"],
                result["counts"]["delete"],
                result["counts"]["error"],
            )
        }] + result["msgs"]

        print("[INFO] total: %s, update: %s, insert: %s, skip: %s, delete: %s, error: %s" % (
            result["counts"]["total"],
            result["counts"]["update"],
            result["counts"]["insert"],
            result["counts"]["skip"],
            result["counts"]["delete"],
            result["counts"]["error"],
        ))

        return result

    # ---
    # function.create
    # ---

    def create_info(self, new_course_info):
        result = {
            "msgs": [],
        }

        # 获取 新增信息
        insert_info = self.new_doc

        course_info_field_headers = self.field_headers
        course_info_field_limits  = self.field_limits
        course_info_field_fixeds  = self.field_fixeds

        default_doc = self.new_doc

        for field in new_course_info:
            new_value = new_course_info[field]
            # 校验字段 固定字段 跳过
            if field in course_info_field_fixeds:
                continue
            # 校验字段 所有字段 默认值
            if not new_value:
                new_value = default_doc[field]
            # 校验字段 特定字段 课程代码
            if field == "course_code" and new_value:
                course_num = mongo.coll_course_info.count_documents({"course_code": new_value})
                if course_num != 0:
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值已存在 [ %s=%s ]" % (course_info_field_headers[field], new_value)
                    })
                    return result
            # 校验字段 特定字段 活动规则
            if field == "activity_rules":
                try:
                    new_value = json.loads(str(new_value).replace("'", '"'))
                except json.JSONDecodeError:
                    result["msgs"].append({
                        "type": "error",
                        "text": "校验失败<br>值不正确 [ %s=%s ]" % (course_info_field_headers[field], new_value)
                    })
                    return result
            # 校验字段 限制字段
            if field in course_info_field_limits and new_value not in course_info_field_limits[field]:
                result["msgs"].append({
                    "type": "error",
                    "text": "校验失败<br>值不可选 [ %s=%s ]" % (course_info_field_headers[field], new_value)
                })
                return result
            # 新增字段
            insert_info[field] = new_value

        # 新增信息
        insert_info["modifytime"] = datetime.datetime.now()
        insert_info["modifyuser"] = current_user.idno
        mongo.coll_course_info.insert_one(insert_info)
        result["msgs"].append({
            "type": "success",
            "text": "新增成功<br>活动信息 [ 总数=1, 新增=1, 错误=0 ]"
        })
        return result

    # ---
    # function.save
    # ---

    def save_infos(self, course_infos):
        result = {
            "msgs": [],
            "file_info": None,
        }

        file_dir = "temp_dir"
        date_info = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = "course_utils-save_infos-%s.xlsx" % date_info
        save_name = "课程信息-%s.xlsx" % date_info
        file_path = file_utils.get_path(file_dir=file_dir, file_name=file_name)

        try:
            wb = Workbook()
            ws = wb.worksheets[0]

            header_fields = {header: field for field, header in self.field_headers.items()}

            alignment = Alignment(
                vertical="center",
                horizontal="center"
            )

            aligntments = {header: alignment for header in header_fields}

            fonts = {
                "header": Font(
                    bold=True
                )
            }

            for col, _ in enumerate(header_fields):
                ws.column_dimensions[get_column_letter(col+1)].width = 20

            for col, header in enumerate(header_fields, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.alignment = aligntments[header]
                cell.font = fonts["header"]
                cell.number_format = numbers.FORMAT_TEXT

            for row, course_info in enumerate(course_infos, start=2):
                for col, header in enumerate(header_fields, start=1):
                    field = header_fields[header]
                    cell = ws.cell(row=row, column=col)
                    cell.value = str(course_info[field]).strip().replace("'", '"')
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.alignment = aligntments[header]

            wb.save(file_path)

        except Exception as e:
            print(traceback.format_exc())
            result["msgs"].append({
                "type": "error",
                "text": "保存文件失败<br>[ file_dir=%s, file_name=%s ]" % (file_dir, file_name)
            })
            return result

        result["file_info"] = {
            "file_dir": file_dir,
            "file_name": file_name,
            "save_name": save_name
        }
        return result


course_utils = CourseUtils()

